"""Manual listings import service for WBR v2 child ASIN catalog."""

from __future__ import annotations

import csv
import io
import os
import re
from dataclasses import dataclass
from datetime import UTC, datetime
from typing import Any

import httpx
from openpyxl import load_workbook
from supabase import Client

from .profiles import WBRNotFoundError, WBRValidationError

LISTING_ALLOWED_EXTENSIONS = (".txt", ".tsv", ".csv", ".xlsx", ".xlsm")
UTF16_BOMS = (b"\xff\xfe", b"\xfe\xff")
WINDSOR_LISTING_FIELDS = [
    "account_id",
    "marketplace_country",
    "merchant_listings_all_data__add_delete",
    "merchant_listings_all_data__asin1",
    "merchant_listings_all_data__asin2",
    "merchant_listings_all_data__asin3",
    "merchant_listings_all_data__bid_for_featured_placement",
    "merchant_listings_all_data__expedited_shipping",
    "merchant_listings_all_data__fulfillment_channel",
    "merchant_listings_all_data__image_url",
    "merchant_listings_all_data__item_condition",
    "merchant_listings_all_data__item_description",
    "merchant_listings_all_data__item_is_marketplace",
    "merchant_listings_all_data__item_name",
    "merchant_listings_all_data__item_note",
    "merchant_listings_all_data__listing_id",
    "merchant_listings_all_data__merchant_shipping_group",
    "merchant_listings_all_data__open_date",
    "merchant_listings_all_data__pending_quantity",
    "merchant_listings_all_data__price",
    "merchant_listings_all_data__product_id",
    "merchant_listings_all_data__product_id_type",
    "merchant_listings_all_data__quantity",
    "merchant_listings_all_data__seller_sku",
    "merchant_listings_all_data__status",
    "merchant_listings_all_data__will_ship_internationally",
    "merchant_listings_all_data__zshop_boldface",
    "merchant_listings_all_data__zshop_browse_path",
    "merchant_listings_all_data__zshop_category1",
    "merchant_listings_all_data__zshop_shipping_fee",
    "merchant_listings_all_data__zshop_storefront_feature",
]
DEFAULT_WINDSOR_LISTING_DATE_PRESET = "last_3d"
DEFAULT_TIMEOUT_SECONDS = 360
MIN_TIMEOUT_SECONDS = 60

LISTING_HEADER_ALIASES = {
    "child_asin": {
        "childasin",
        "asin",
        "asin1",
        "merchantlistingsalldataasin1",
    },
    "parent_asin": {
        "parentasin",
        "merchantlistingsalldataasin2",
    },
    "child_sku": {
        "childsku",
        "sellersku",
        "sku",
        "merchantlistingsalldatasellersku",
    },
    "parent_sku": {
        "parentsku",
    },
    "child_product_name": {
        "childproductname",
        "itemname",
        "productname",
        "merchantlistingsalldataitemname",
    },
    "parent_title": {
        "parenttitle",
        "title",
    },
    "fnsku": {
        "fnsku",
    },
    "upc": {
        "upc",
    },
    "category": {
        "category",
        "zshopcategory1",
        "merchantlistingsalldatazshopcategory1",
    },
    "source_item_style": {
        "itemstyle",
        "style",
        "sourceitemstyle",
    },
    "size": {
        "size",
    },
    "fulfillment_method": {
        "fulfillmentmethod",
        "fulfillmentchannel",
        "merchantlistingsalldatafulfillmentchannel",
    },
    "product_id": {
        "productid",
        "merchantlistingsalldataproductid",
    },
    "product_id_type": {
        "productidtype",
        "merchantlistingsalldataproductidtype",
    },
}

FOOTER_LABELS = {"total", "total:"}
ASIN_PATTERN = re.compile(r"^[A-Z0-9]{10}$")


@dataclass(frozen=True)
class ParsedListingRecord:
    child_asin: str
    parent_asin: str | None
    parent_sku: str | None
    child_sku: str | None
    fnsku: str | None
    upc: str | None
    category: str | None
    parent_title: str | None
    child_product_name: str | None
    source_item_style: str | None
    size: str | None
    fulfillment_method: str | None
    raw_payload: dict[str, Any]


@dataclass(frozen=True)
class ParsedListingFile:
    source_type: str
    sheet_title: str | None
    header_row_index: int
    rows_read: int
    duplicate_rows_merged: int
    records: list[ParsedListingRecord]


def _canonicalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def _as_cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_optional(value: str) -> str | None:
    cleaned = value.strip()
    return cleaned or None


def _is_footer_row(value: str) -> bool:
    return value.strip().lower() in FOOTER_LABELS


def _looks_like_asin(value: str | None) -> bool:
    if not value:
        return False
    return bool(ASIN_PATTERN.fullmatch(value.strip().upper()))


def _pick_row_value(row: tuple[Any, ...] | list[str], header_map: dict[str, int], field_name: str) -> str:
    idx = header_map.get(field_name)
    if idx is None or idx >= len(row):
        return ""
    return _as_cell_text(row[idx])


def _map_headers(header_values: list[str]) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for idx, value in enumerate(header_values):
        canonical = _canonicalize_header(value)
        if not canonical:
            continue
        for field_name, aliases in LISTING_HEADER_ALIASES.items():
            if canonical in aliases and field_name not in header_map:
                header_map[field_name] = idx
    return header_map


def _parse_record_from_row(
    row: tuple[Any, ...] | list[str],
    header_values: list[str],
    header_map: dict[str, int],
) -> ParsedListingRecord | None:
    child_asin = _pick_row_value(row, header_map, "child_asin").upper()
    product_id = _pick_row_value(row, header_map, "product_id").upper()
    product_id_type = _pick_row_value(row, header_map, "product_id_type").upper()
    if not child_asin and product_id_type == "ASIN" and _looks_like_asin(product_id):
        child_asin = product_id

    child_sku = _pick_row_value(row, header_map, "child_sku")
    child_product_name = _pick_row_value(row, header_map, "child_product_name")

    if not child_asin:
        leading_value = child_sku or child_product_name or _as_cell_text(row[0]) if row else ""
        if not leading_value or _is_footer_row(leading_value):
            return None
        raise WBRValidationError("Listings rows must include a child ASIN (or product-id ASIN fallback)")

    if not _looks_like_asin(child_asin):
        raise WBRValidationError(f'Child ASIN "{child_asin}" is not a valid Amazon ASIN')

    raw_payload = {
        header_values[idx]: _as_cell_text(row[idx]) if idx < len(row) else ""
        for idx in range(len(header_values))
        if header_values[idx]
    }

    return ParsedListingRecord(
        child_asin=child_asin,
        parent_asin=_normalize_optional(_pick_row_value(row, header_map, "parent_asin")),
        parent_sku=_normalize_optional(_pick_row_value(row, header_map, "parent_sku")),
        child_sku=_normalize_optional(child_sku),
        fnsku=_normalize_optional(_pick_row_value(row, header_map, "fnsku")),
        upc=_normalize_optional(_pick_row_value(row, header_map, "upc")),
        category=_normalize_optional(_pick_row_value(row, header_map, "category")),
        parent_title=_normalize_optional(_pick_row_value(row, header_map, "parent_title")),
        child_product_name=_normalize_optional(child_product_name),
        source_item_style=_normalize_optional(_pick_row_value(row, header_map, "source_item_style")),
        size=_normalize_optional(_pick_row_value(row, header_map, "size")),
        fulfillment_method=_normalize_optional(_pick_row_value(row, header_map, "fulfillment_method")),
        raw_payload=raw_payload,
    )


def _merge_record(existing: ParsedListingRecord, incoming: ParsedListingRecord) -> ParsedListingRecord:
    return ParsedListingRecord(
        child_asin=existing.child_asin,
        parent_asin=incoming.parent_asin or existing.parent_asin,
        parent_sku=incoming.parent_sku or existing.parent_sku,
        child_sku=incoming.child_sku or existing.child_sku,
        fnsku=incoming.fnsku or existing.fnsku,
        upc=incoming.upc or existing.upc,
        category=incoming.category or existing.category,
        parent_title=incoming.parent_title or existing.parent_title,
        child_product_name=incoming.child_product_name or existing.child_product_name,
        source_item_style=incoming.source_item_style or existing.source_item_style,
        size=incoming.size or existing.size,
        fulfillment_method=incoming.fulfillment_method or existing.fulfillment_method,
        raw_payload={**existing.raw_payload, **incoming.raw_payload},
    )


def _parse_sheet_rows(rows: list[tuple[Any, ...]], *, source_type: str, sheet_title: str | None) -> ParsedListingFile:
    header_row_index = -1
    header_values: list[str] = []
    header_map: dict[str, int] = {}

    for idx, row in enumerate(rows):
        values = [_as_cell_text(cell) for cell in row]
        mapped_headers = _map_headers(values)
        if "child_asin" in mapped_headers or ("product_id" in mapped_headers and "product_id_type" in mapped_headers):
            header_row_index = idx
            header_values = values
            header_map = mapped_headers
            break

    if header_row_index < 0:
        raise WBRValidationError("Listings file must include a child ASIN column such as Child ASIN or asin1")

    deduped_by_asin: dict[str, ParsedListingRecord] = {}
    ordered_records: list[ParsedListingRecord] = []
    duplicate_rows_merged = 0

    for row in rows[header_row_index + 1 :]:
        parsed = _parse_record_from_row(row, header_values, header_map)
        if parsed is None:
            continue
        existing = deduped_by_asin.get(parsed.child_asin)
        if existing is None:
            deduped_by_asin[parsed.child_asin] = parsed
            ordered_records.append(parsed)
            continue
        merged = _merge_record(existing, parsed)
        deduped_by_asin[parsed.child_asin] = merged
        duplicate_rows_merged += 1
        for idx, current in enumerate(ordered_records):
            if current.child_asin == parsed.child_asin:
                ordered_records[idx] = merged
                break

    if not ordered_records:
        raise WBRValidationError("Listings file contained no child ASIN rows")

    return ParsedListingFile(
        source_type=source_type,
        sheet_title=sheet_title,
        header_row_index=header_row_index,
        rows_read=len(ordered_records) + duplicate_rows_merged,
        duplicate_rows_merged=duplicate_rows_merged,
        records=ordered_records,
    )


def _parse_dict_rows(
    rows: list[dict[str, Any]], *, source_type: str, sheet_title: str | None
) -> ParsedListingFile:
    if not rows:
        raise WBRValidationError("Listings source returned no rows")

    header_values = list(rows[0].keys())
    header_map = _map_headers(header_values)
    if "child_asin" not in header_map and not (
        "product_id" in header_map and "product_id_type" in header_map
    ):
        raise WBRValidationError("Listings source must include a child ASIN column such as Child ASIN or asin1")

    deduped_by_asin: dict[str, ParsedListingRecord] = {}
    ordered_records: list[ParsedListingRecord] = []
    duplicate_rows_merged = 0

    for row in rows:
        tuple_row = tuple(row.get(header) for header in header_values)
        parsed = _parse_record_from_row(tuple_row, header_values, header_map)
        if parsed is None:
            continue
        existing = deduped_by_asin.get(parsed.child_asin)
        if existing is None:
            deduped_by_asin[parsed.child_asin] = parsed
            ordered_records.append(parsed)
            continue
        merged = _merge_record(existing, parsed)
        deduped_by_asin[parsed.child_asin] = merged
        duplicate_rows_merged += 1
        for idx, current in enumerate(ordered_records):
            if current.child_asin == parsed.child_asin:
                ordered_records[idx] = merged
                break

    if not ordered_records:
        raise WBRValidationError("Listings source contained no child ASIN rows")

    return ParsedListingFile(
        source_type=source_type,
        sheet_title=sheet_title,
        header_row_index=0,
        rows_read=len(ordered_records) + duplicate_rows_merged,
        duplicate_rows_merged=duplicate_rows_merged,
        records=ordered_records,
    )


def _parse_spreadsheet(file_bytes: bytes) -> ParsedListingFile:
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise WBRValidationError(f"Unable to read listings workbook: {exc}") from exc

    for sheet in workbook.worksheets:
        sheet.reset_dimensions()
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        try:
            return _parse_sheet_rows(rows, source_type="spreadsheet", sheet_title=sheet.title)
        except WBRValidationError:
            continue

    raise WBRValidationError("Listings workbook must include a child ASIN column such as Child ASIN or asin1")


def _parse_delimited(file_bytes: bytes) -> ParsedListingFile:
    text: str | None = None
    if file_bytes.startswith(UTF16_BOMS):
        try:
            text = file_bytes.decode("utf-16")
        except UnicodeDecodeError as exc:
            raise WBRValidationError("Unable to decode UTF-16 listings file") from exc
    else:
        for encoding in ("utf-8-sig", "cp1252", "latin-1"):
            try:
                text = file_bytes.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
    if text is None:
        raise WBRValidationError("Unable to decode listings file")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters="\t,;|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = "\t" if "\t" in sample else ","

    rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    if not rows:
        raise WBRValidationError("Listings file is empty")
    return _parse_sheet_rows([tuple(row) for row in rows], source_type="delimited", sheet_title=None)


def parse_listing_file(file_name: str, file_bytes: bytes) -> ParsedListingFile:
    lower_name = (file_name or "").lower()
    if not lower_name.endswith(LISTING_ALLOWED_EXTENSIONS):
        raise WBRValidationError("Listings import supports .txt, .tsv, .csv, .xlsx, and .xlsm files only")

    if lower_name.endswith((".xlsx", ".xlsm")):
        return _parse_spreadsheet(file_bytes)
    return _parse_delimited(file_bytes)


class ListingImportService:
    def __init__(self, db: Client) -> None:
        self.db = db
        self.windsor_api_key = os.getenv("WINDSOR_API_KEY", "").strip()
        self.windsor_seller_url = os.getenv("WINDSOR_SELLER_URL", "https://connectors.windsor.ai/amazon_sp").strip()
        self.windsor_listing_date_preset = os.getenv(
            "WBR_WINDSOR_LISTING_DATE_PRESET", DEFAULT_WINDSOR_LISTING_DATE_PRESET
        ).strip() or DEFAULT_WINDSOR_LISTING_DATE_PRESET
        configured_timeout = int(os.getenv("WBR_REQUEST_TIMEOUT_SECONDS", str(DEFAULT_TIMEOUT_SECONDS)))
        self.timeout_seconds = max(configured_timeout, MIN_TIMEOUT_SECONDS)

    def list_import_batches(self, profile_id: str) -> list[dict[str, Any]]:
        self._get_profile(profile_id)
        response = (
            self.db.table("wbr_listing_import_batches")
            .select("*")
            .eq("profile_id", profile_id)
            .order("created_at", desc=True)
            .limit(20)
            .execute()
        )
        return response.data if isinstance(response.data, list) else []

    async def import_from_windsor(
        self,
        *,
        profile_id: str,
        user_id: str | None = None,
    ) -> dict[str, Any]:
        profile = self._get_profile(profile_id)
        account_id = str(profile.get("windsor_account_id") or "").strip()
        if not account_id:
            raise WBRValidationError("WBR profile is missing windsor_account_id")
        if not self.windsor_api_key:
            raise WBRValidationError("WINDSOR_API_KEY is not configured")
        if not self.windsor_seller_url:
            raise WBRValidationError("WINDSOR_SELLER_URL is not configured")

        batch = self._create_batch(
            profile_id=profile_id,
            file_name=f"windsor:get_merchant_listings_all_data:{account_id}",
            user_id=user_id,
        )
        batch_id = str(batch["id"])

        try:
            rows = await self._fetch_windsor_rows(account_id=account_id)
            self._validate_marketplace(rows, expected_marketplace=str(profile.get("marketplace_code") or "").strip())
            parsed = _parse_dict_rows(rows, source_type="windsor", sheet_title=None)
            loaded_count = self._replace_child_asins(
                profile_id=profile_id,
                batch_id=batch_id,
                records=parsed.records,
            )
            batch = self._finalize_batch(
                batch_id=batch_id,
                import_status="success",
                rows_read=parsed.rows_read,
                rows_loaded=loaded_count,
                error_message=None,
            )
            return {
                "batch": batch,
                "summary": {
                    "source_type": parsed.source_type,
                    "sheet_title": parsed.sheet_title,
                    "header_row_index": parsed.header_row_index,
                    "rows_read": parsed.rows_read,
                    "rows_loaded": loaded_count,
                    "duplicate_rows_merged": parsed.duplicate_rows_merged,
                    "windsor_account_id": account_id,
                },
            }
        except WBRValidationError as exc:
            self._finalize_batch(
                batch_id=batch_id,
                import_status="error",
                rows_read=0,
                rows_loaded=0,
                error_message=str(exc),
            )
            raise
        except Exception as exc:  # noqa: BLE001
            self._finalize_batch(
                batch_id=batch_id,
                import_status="error",
                rows_read=0,
                rows_loaded=0,
                error_message=str(exc),
            )
            raise

    def import_file(
        self,
        *,
        profile_id: str,
        file_name: str,
        file_bytes: bytes,
        user_id: str | None = None,
    ) -> dict[str, Any]:
        self._validate_file_name(file_name)
        self._get_profile(profile_id)
        batch = self._create_batch(profile_id=profile_id, file_name=file_name, user_id=user_id)
        batch_id = str(batch["id"])

        try:
            parsed = parse_listing_file(file_name, file_bytes)
            loaded_count = self._replace_child_asins(
                profile_id=profile_id,
                batch_id=batch_id,
                records=parsed.records,
            )
            batch = self._finalize_batch(
                batch_id=batch_id,
                import_status="success",
                rows_read=parsed.rows_read,
                rows_loaded=loaded_count,
                error_message=None,
            )
            return {
                "batch": batch,
                "summary": {
                    "source_type": parsed.source_type,
                    "sheet_title": parsed.sheet_title,
                    "header_row_index": parsed.header_row_index,
                    "rows_read": parsed.rows_read,
                    "rows_loaded": loaded_count,
                    "duplicate_rows_merged": parsed.duplicate_rows_merged,
                },
            }
        except WBRValidationError as exc:
            self._finalize_batch(
                batch_id=batch_id,
                import_status="error",
                rows_read=0,
                rows_loaded=0,
                error_message=str(exc),
            )
            raise
        except Exception as exc:  # noqa: BLE001
            self._finalize_batch(
                batch_id=batch_id,
                import_status="error",
                rows_read=0,
                rows_loaded=0,
                error_message=str(exc),
            )
            raise

    def _validate_file_name(self, file_name: str) -> None:
        lower_name = (file_name or "").lower()
        if not lower_name.endswith(LISTING_ALLOWED_EXTENSIONS):
            raise WBRValidationError("Listings import supports .txt, .tsv, .csv, .xlsx, and .xlsm files only")

    def _get_profile(self, profile_id: str) -> dict[str, Any]:
        response = (
            self.db.table("wbr_profiles")
            .select("*")
            .eq("id", profile_id)
            .limit(1)
            .execute()
        )
        rows = response.data if isinstance(response.data, list) else []
        if not rows:
            raise WBRNotFoundError(f"Profile {profile_id} not found")
        return rows[0]

    async def _fetch_windsor_rows(self, *, account_id: str) -> list[dict[str, Any]]:
        params = {
            "api_key": self.windsor_api_key,
            "date_preset": self.windsor_listing_date_preset,
            "fields": ",".join(WINDSOR_LISTING_FIELDS),
            "select_accounts": account_id,
        }
        response = await self._request_windsor(params)
        if response.status_code >= 400:
            body_preview = response.text.strip().replace("\n", " ")[:220]
            raise WBRValidationError(
                f"Windsor listings request failed: {response.status_code} :: {body_preview}"
            )
        return self._parse_windsor_response_rows(response)

    async def _request_windsor(self, params: dict[str, str]) -> httpx.Response:
        timeout = httpx.Timeout(timeout=self.timeout_seconds)
        async with httpx.AsyncClient(timeout=timeout) as client:
            return await client.get(self.windsor_seller_url, params=params)

    def _parse_windsor_response_rows(self, response: httpx.Response) -> list[dict[str, Any]]:
        body = response.text.strip()
        if not body:
            return []

        content_type = response.headers.get("content-type", "").lower()
        if "json" in content_type or body.startswith("[") or body.startswith("{"):
            payload = response.json()
            if isinstance(payload, list):
                return [row for row in payload if isinstance(row, dict)]
            if isinstance(payload, dict):
                for key in ("data", "results"):
                    value = payload.get(key)
                    if isinstance(value, list):
                        return [row for row in value if isinstance(row, dict)]
            raise WBRValidationError("Unexpected Windsor listings payload shape")

        reader = csv.DictReader(io.StringIO(body))
        return [dict(row) for row in reader]

    def _validate_marketplace(self, rows: list[dict[str, Any]], *, expected_marketplace: str) -> None:
        expected = expected_marketplace.strip().upper()
        if not expected:
            return
        marketplace_values = {
            str(row.get("marketplace_country") or "").strip().upper()
            for row in rows
            if str(row.get("marketplace_country") or "").strip()
        }
        if not marketplace_values:
            return
        if marketplace_values != {expected}:
            raise WBRValidationError(
                f"Windsor listings marketplace mismatch: expected {expected}, got {', '.join(sorted(marketplace_values))}"
            )

    def _create_batch(self, *, profile_id: str, file_name: str, user_id: str | None) -> dict[str, Any]:
        payload: dict[str, Any] = {
            "profile_id": profile_id,
            "source_filename": file_name,
            "import_status": "running",
        }
        if user_id:
            payload["initiated_by"] = user_id
        response = self.db.table("wbr_listing_import_batches").insert(payload).execute()
        rows = response.data if isinstance(response.data, list) else []
        if not rows:
            raise WBRValidationError("Failed to create listings import batch")
        return rows[0]

    def _finalize_batch(
        self,
        *,
        batch_id: str,
        import_status: str,
        rows_read: int,
        rows_loaded: int,
        error_message: str | None,
    ) -> dict[str, Any]:
        response = (
            self.db.table("wbr_listing_import_batches")
            .update(
                {
                    "import_status": import_status,
                    "rows_read": rows_read,
                    "rows_loaded": rows_loaded,
                    "error_message": error_message,
                    "finished_at": datetime.now(UTC).isoformat(),
                }
            )
            .eq("id", batch_id)
            .execute()
        )
        rows = response.data if isinstance(response.data, list) else []
        if not rows:
            raise WBRValidationError("Failed to finalize listings import batch")
        return rows[0]

    def _replace_child_asins(
        self,
        *,
        profile_id: str,
        batch_id: str,
        records: list[ParsedListingRecord],
    ) -> int:
        (
            self.db.table("wbr_profile_child_asins")
            .update({"active": False})
            .eq("profile_id", profile_id)
            .eq("active", True)
            .execute()
        )

        payloads = [
            {
                "profile_id": profile_id,
                "listing_batch_id": batch_id,
                "parent_asin": record.parent_asin,
                "child_asin": record.child_asin,
                "parent_sku": record.parent_sku,
                "child_sku": record.child_sku,
                "fnsku": record.fnsku,
                "upc": record.upc,
                "category": record.category,
                "parent_title": record.parent_title,
                "child_product_name": record.child_product_name,
                "source_item_style": record.source_item_style,
                "size": record.size,
                "fulfillment_method": record.fulfillment_method,
                "raw_payload": record.raw_payload,
                "active": True,
            }
            for record in records
        ]

        for start in range(0, len(payloads), 500):
            chunk = payloads[start : start + 500]
            response = self.db.table("wbr_profile_child_asins").insert(chunk).execute()
            rows = response.data if isinstance(response.data, list) else []
            if len(rows) != len(chunk):
                raise WBRValidationError("Failed to store imported child ASIN catalog rows")

        return len(payloads)
