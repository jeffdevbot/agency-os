"""Pacvue import service for WBR v2 campaign mappings."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import datetime, UTC
from typing import Any

from openpyxl import load_workbook
from postgrest.exceptions import APIError as PostgrestAPIError
from supabase import Client

from .profiles import WBRNotFoundError, WBRValidationError

PACVUE_REQUIRED_HEADERS = {
    "name": "campaign_name",
    "campaigntagnames": "raw_tag",
}
PACVUE_ALLOWED_EXTENSIONS = (".xlsx", ".xlsm")
PACVUE_GOAL_CODES = {
    "perf": "Perf",
    "rsrch": "Rsrch",
    "comp": "Comp",
    "category": "Comp",
    "competitor": "Comp",
    "harv": "Harv",
    "def": "Def",
    "rank": "Rank",
}
PACVUE_EMPTY_TAG_PATTERN = re.compile(r"^[-\s\u2013\u2014]+$")
PACVUE_FOOTER_LABELS = {"total", "total:"}

_BATCH_SIZE = 500


@dataclass(frozen=True)
class ParsedPacvueRecord:
    campaign_name: str
    raw_tag: str
    leaf_row_label: str
    goal_code: str
    raw_payload: dict[str, Any]


def _is_zero_metric_value(value: Any) -> bool:
    text = _as_cell_text(value)
    if not text:
        return True
    normalized = text.replace(",", "")
    try:
        return float(normalized) == 0.0
    except ValueError:
        return False


def _record_is_archived_zero_duplicate(record: ParsedPacvueRecord) -> bool:
    state = _as_cell_text(record.raw_payload.get("state")).lower()
    if state != "archived":
        return False
    for key in ("Impression", "Click", "Spend", "Sales", "Orders"):
        if not _is_zero_metric_value(record.raw_payload.get(key)):
            return False
    return True


def _prefer_pacvue_record(existing: ParsedPacvueRecord, candidate: ParsedPacvueRecord) -> ParsedPacvueRecord | None:
    if existing.raw_tag == candidate.raw_tag:
        return existing
    existing_archived_zero = _record_is_archived_zero_duplicate(existing)
    candidate_archived_zero = _record_is_archived_zero_duplicate(candidate)
    if existing_archived_zero and not candidate_archived_zero:
        return candidate
    if candidate_archived_zero and not existing_archived_zero:
        return existing
    return None


@dataclass(frozen=True)
class ParsedPacvueWorkbook:
    sheet_title: str
    header_row_index: int
    rows_read: int
    duplicate_rows_skipped: int
    unmapped_rows_skipped: int
    invalid_rows_skipped: int
    warnings: list[str]
    records: list[ParsedPacvueRecord]


def _canonicalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def _as_cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_pacvue_workbook(file_bytes: bytes) -> ParsedPacvueWorkbook:
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise WBRValidationError(f"Unable to read Pacvue workbook: {exc}") from exc

    selected_sheet_title: str | None = None
    rows: list[tuple[Any, ...]] | None = None
    header_row_index: int | None = None
    header_map: dict[str, int] | None = None
    header_values: list[str] | None = None

    for sheet in workbook.worksheets:
        # Pacvue exports can declare an incorrect worksheet dimension like A1
        # even when the sheet contains a full table. Resetting dimensions lets
        # openpyxl scan the actual populated cells instead of trusting the bad
        # metadata.
        sheet.reset_dimensions()
        sheet_rows = list(sheet.iter_rows(values_only=True))
        if not sheet_rows:
            continue
        try:
            found_header_row_index, found_header_map, found_header_values = _find_pacvue_header(sheet_rows)
        except WBRValidationError:
            continue

        selected_sheet_title = sheet.title
        rows = sheet_rows
        header_row_index = found_header_row_index
        header_map = found_header_map
        header_values = found_header_values
        break

    if rows is None or header_row_index is None or header_map is None or header_values is None:
        raise WBRValidationError('Pacvue workbook must contain "Name" and "CampaignTagNames" columns')

    records: list[ParsedPacvueRecord] = []
    deduped_by_campaign: dict[str, ParsedPacvueRecord] = {}
    duplicate_rows_skipped = 0
    unmapped_rows_skipped = 0
    invalid_rows_skipped = 0
    warning_counts: dict[str, int] = {}

    for row in rows[header_row_index + 1 :]:
        campaign_name = _as_cell_text(row[header_map["campaign_name"]]) if header_map["campaign_name"] < len(row) else ""
        raw_tag = _as_cell_text(row[header_map["raw_tag"]]) if header_map["raw_tag"] < len(row) else ""

        # Ignore fully blank rows beneath the header.
        if not campaign_name and not raw_tag:
            continue
        if campaign_name and not raw_tag and _is_footer_row(campaign_name):
            continue
        if not campaign_name or not raw_tag:
            invalid_rows_skipped += 1
            warning_counts["Pacvue rows must include both Name and CampaignTagNames values"] = (
                warning_counts.get("Pacvue rows must include both Name and CampaignTagNames values", 0) + 1
            )
            continue
        if _is_empty_tag(raw_tag):
            unmapped_rows_skipped += 1
            continue

        try:
            leaf_row_label, goal_code = _parse_tag(raw_tag)
        except WBRValidationError as exc:
            invalid_rows_skipped += 1
            warning_counts[str(exc)] = warning_counts.get(str(exc), 0) + 1
            continue
        raw_payload = {
            header_values[idx]: _as_cell_text(row[idx]) if idx < len(row) else ""
            for idx in range(len(header_values))
            if header_values[idx]
        }
        record = ParsedPacvueRecord(
            campaign_name=campaign_name,
            raw_tag=raw_tag,
            leaf_row_label=leaf_row_label,
            goal_code=goal_code,
            raw_payload=raw_payload,
        )

        existing = deduped_by_campaign.get(campaign_name)
        if existing is None:
            deduped_by_campaign[campaign_name] = record
            records.append(record)
            continue
        preferred = _prefer_pacvue_record(existing, record)
        if preferred is None:
            raise WBRValidationError(
                f'Campaign "{campaign_name}" appears multiple times with conflicting Pacvue tags'
            )
        if preferred is not existing:
            deduped_by_campaign[campaign_name] = preferred
            record_index = records.index(existing)
            records[record_index] = preferred
        duplicate_rows_skipped += 1

    if not records:
        if warning_counts:
            top_message = sorted(warning_counts.items(), key=lambda item: (-item[1], item[0]))[0][0]
            raise WBRValidationError(top_message)
        raise WBRValidationError("Pacvue workbook contained no campaign/tag rows")

    warnings = [
        f"{message} ({count} row{'s' if count != 1 else ''})"
        for message, count in sorted(warning_counts.items(), key=lambda item: (-item[1], item[0]))
    ]

    return ParsedPacvueWorkbook(
        sheet_title=selected_sheet_title or workbook.active.title,
        header_row_index=header_row_index,
        rows_read=len(records) + duplicate_rows_skipped + unmapped_rows_skipped + invalid_rows_skipped,
        duplicate_rows_skipped=duplicate_rows_skipped,
        unmapped_rows_skipped=unmapped_rows_skipped,
        invalid_rows_skipped=invalid_rows_skipped,
        warnings=warnings,
        records=records,
    )


def _find_pacvue_header(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int], list[str]]:
    for idx, row in enumerate(rows):
        normalized = [_canonicalize_header(cell) for cell in row]
        header_positions: dict[str, int] = {}
        for col_idx, value in enumerate(normalized):
            field_name = PACVUE_REQUIRED_HEADERS.get(value)
            if field_name:
                header_positions[field_name] = col_idx
        if len(header_positions) == len(PACVUE_REQUIRED_HEADERS):
            return idx, header_positions, [_as_cell_text(cell) for cell in row]
    raise WBRValidationError('Pacvue workbook must contain "Name" and "CampaignTagNames" columns')


def _parse_tag(raw_tag: str) -> tuple[str, str]:
    tag_parts = [part.strip() for part in re.split(r"[\n\r,]+", raw_tag) if part.strip()]
    if len(tag_parts) != 1:
        raise WBRValidationError(
            f'Pacvue tag "{raw_tag}" must contain exactly one tag value for WBR import'
        )

    tag = tag_parts[0]
    if "/" not in tag:
        raise WBRValidationError(f'Pacvue tag "{raw_tag}" is missing a "/ goal" suffix')

    leaf_row_label, goal_code_raw = [part.strip() for part in tag.rsplit("/", 1)]
    if not leaf_row_label:
        raise WBRValidationError(f'Pacvue tag "{raw_tag}" is missing a leaf row label')

    goal_code = PACVUE_GOAL_CODES.get(goal_code_raw.lower())
    if not goal_code:
        raise WBRValidationError(f'Pacvue tag "{raw_tag}" has an unsupported goal suffix')

    return leaf_row_label, goal_code


def _is_empty_tag(raw_tag: str) -> bool:
    return bool(PACVUE_EMPTY_TAG_PATTERN.fullmatch(raw_tag.strip()))


def _is_footer_row(campaign_name: str) -> bool:
    return campaign_name.strip().lower() in PACVUE_FOOTER_LABELS


class PacvueImportService:
    def __init__(self, db: Client) -> None:
        self.db = db

    def list_import_batches(self, profile_id: str) -> list[dict[str, Any]]:
        self._get_profile(profile_id)
        response = (
            self.db.table("wbr_pacvue_import_batches")
            .select("*")
            .eq("profile_id", profile_id)
            .order("created_at", desc=True)
            .limit(20)
            .execute()
        )
        return response.data if isinstance(response.data, list) else []

    def import_workbook(
        self,
        *,
        profile_id: str,
        file_name: str,
        file_bytes: bytes,
        user_id: str | None = None,
    ) -> dict[str, Any]:
        self._get_profile(profile_id)
        self._validate_file_name(file_name)

        batch = self._create_batch(profile_id=profile_id, file_name=file_name, user_id=user_id)
        batch_id = str(batch["id"])

        try:
            parsed = parse_pacvue_workbook(file_bytes)
            row_summary = self._ensure_leaf_rows(profile_id=profile_id, records=parsed.records, user_id=user_id)
            loaded_count = self._refresh_campaign_mappings(
                profile_id=profile_id,
                batch_id=batch_id,
                records=parsed.records,
            )
            batch = self._finalize_batch(
                batch_id=batch_id,
                import_status="success",
                rows_read=parsed.rows_read,
                rows_loaded=loaded_count,
                error_message=self._warning_summary(parsed),
            )
            return {
                "batch": batch,
                "summary": {
                    "header_row_index": parsed.header_row_index,
                    "rows_read": parsed.rows_read,
                    "rows_loaded": loaded_count,
                    "duplicate_rows_skipped": parsed.duplicate_rows_skipped,
                    "unmapped_rows_skipped": parsed.unmapped_rows_skipped,
                    "invalid_rows_skipped": parsed.invalid_rows_skipped,
                    "created_leaf_rows": row_summary["created_leaf_rows"],
                    "reactivated_leaf_rows": row_summary["reactivated_leaf_rows"],
                    "warnings": parsed.warnings,
                },
            }
        except WBRValidationError as exc:
            self._finalize_batch(
                batch_id=batch_id,
                import_status="error",
                rows_read=0,
                rows_loaded=0,
                error_message=str(exc),
            )
            raise
        except Exception as exc:  # noqa: BLE001
            self._finalize_batch(
                batch_id=batch_id,
                import_status="error",
                rows_read=0,
                rows_loaded=0,
                error_message=str(exc),
            )
            raise

    def _validate_file_name(self, file_name: str) -> None:
        lower_name = (file_name or "").lower()
        if not lower_name.endswith(PACVUE_ALLOWED_EXTENSIONS):
            raise WBRValidationError("Pacvue import currently supports .xlsx and .xlsm files only")

    def _warning_summary(self, parsed: ParsedPacvueWorkbook) -> str | None:
        if parsed.invalid_rows_skipped <= 0:
            return None
        top_warning = parsed.warnings[0] if parsed.warnings else "Invalid Pacvue tags were skipped"
        return (
            f"Skipped {parsed.invalid_rows_skipped} invalid Pacvue row"
            f"{'s' if parsed.invalid_rows_skipped != 1 else ''}. "
            f"Those campaigns will remain in Unmapped / Legacy Campaigns until fixed upstream. "
            f"{top_warning}"
        )

    def _get_profile(self, profile_id: str) -> dict[str, Any]:
        response = (
            self.db.table("wbr_profiles")
            .select("*")
            .eq("id", profile_id)
            .limit(1)
            .execute()
        )
        rows = response.data if isinstance(response.data, list) else []
        if not rows:
            raise WBRNotFoundError(f"Profile {profile_id} not found")
        return rows[0]

    def _create_batch(self, *, profile_id: str, file_name: str, user_id: str | None) -> dict[str, Any]:
        payload: dict[str, Any] = {
            "profile_id": profile_id,
            "source_filename": file_name,
            "import_status": "running",
        }
        if user_id:
            payload["initiated_by"] = user_id
        response = self.db.table("wbr_pacvue_import_batches").insert(payload).execute()
        rows = response.data if isinstance(response.data, list) else []
        if not rows:
            raise WBRValidationError("Failed to create Pacvue import batch")
        return rows[0]

    def _finalize_batch(
        self,
        *,
        batch_id: str,
        import_status: str,
        rows_read: int,
        rows_loaded: int,
        error_message: str | None,
    ) -> dict[str, Any]:
        response = (
            self.db.table("wbr_pacvue_import_batches")
            .update(
                {
                    "import_status": import_status,
                    "rows_read": rows_read,
                    "rows_loaded": rows_loaded,
                    "error_message": error_message,
                    "finished_at": datetime.now(UTC).isoformat(),
                }
            )
            .eq("id", batch_id)
            .execute()
        )
        rows = response.data if isinstance(response.data, list) else []
        if not rows:
            raise WBRValidationError("Failed to finalize Pacvue import batch")
        return rows[0]

    def _ensure_leaf_rows(
        self,
        *,
        profile_id: str,
        records: list[ParsedPacvueRecord],
        user_id: str | None,
    ) -> dict[str, int]:
        desired_labels: list[str] = []
        seen_labels: set[str] = set()
        for record in records:
            if record.leaf_row_label in seen_labels:
                continue
            desired_labels.append(record.leaf_row_label)
            seen_labels.add(record.leaf_row_label)

        existing_rows_response = (
            self.db.table("wbr_rows")
            .select("*")
            .eq("profile_id", profile_id)
            .eq("row_kind", "leaf")
            .execute()
        )
        existing_rows = existing_rows_response.data if isinstance(existing_rows_response.data, list) else []
        existing_by_label = {str(row["row_label"]): row for row in existing_rows if isinstance(row, dict)}

        max_sort_order = max((int(row.get("sort_order") or 0) for row in existing_rows), default=0)
        created_leaf_rows = 0
        reactivated_leaf_rows = 0

        for label in desired_labels:
            existing = existing_by_label.get(label)
            if existing:
                if not existing.get("active"):
                    updates: dict[str, Any] = {"active": True}
                    if user_id:
                        updates["updated_by"] = user_id
                    (
                        self.db.table("wbr_rows")
                        .update(updates)
                        .eq("id", existing["id"])
                        .execute()
                    )
                    reactivated_leaf_rows += 1
                continue

            max_sort_order += 10
            payload: dict[str, Any] = {
                "profile_id": profile_id,
                "row_label": label,
                "row_kind": "leaf",
                "sort_order": max_sort_order,
            }
            if user_id:
                payload["created_by"] = user_id
                payload["updated_by"] = user_id
            try:
                response = self.db.table("wbr_rows").insert(payload).execute()
            except PostgrestAPIError as exc:
                raise WBRValidationError(str(exc)) from exc
            rows = response.data if isinstance(response.data, list) else []
            if not rows:
                raise WBRValidationError(f'Failed to create leaf row "{label}" during Pacvue import')
            created_leaf_rows += 1

        return {
            "created_leaf_rows": created_leaf_rows,
            "reactivated_leaf_rows": reactivated_leaf_rows,
        }

    def _refresh_campaign_mappings(
        self,
        *,
        profile_id: str,
        batch_id: str,
        records: list[ParsedPacvueRecord],
    ) -> int:
        active_leaf_rows_response = (
            self.db.table("wbr_rows")
            .select("id,row_label")
            .eq("profile_id", profile_id)
            .eq("row_kind", "leaf")
            .eq("active", True)
            .execute()
        )
        active_leaf_rows = active_leaf_rows_response.data if isinstance(active_leaf_rows_response.data, list) else []
        row_id_by_label = {
            str(row["row_label"]): str(row["id"])
            for row in active_leaf_rows
            if isinstance(row, dict)
        }

        insert_payloads: list[dict[str, Any]] = []
        for record in records:
            row_id = row_id_by_label.get(record.leaf_row_label)
            if not row_id:
                raise WBRValidationError(
                    f'Leaf row "{record.leaf_row_label}" was not available after Pacvue import row setup'
                )
            insert_payloads.append(
                {
                    "profile_id": profile_id,
                    "import_batch_id": batch_id,
                    "campaign_name": record.campaign_name,
                    "raw_tag": record.raw_tag,
                    "row_id": row_id,
                    "leaf_row_label": record.leaf_row_label,
                    "goal_code": record.goal_code,
                    "raw_payload": record.raw_payload,
                    "active": False,
                }
            )

        try:
            self.db.table("wbr_pacvue_campaign_map").insert(insert_payloads).execute()
        except PostgrestAPIError as exc:
            raise WBRValidationError(str(exc)) from exc

        # Only deactivate prior mappings for campaigns the new batch actually
        # tagged. Campaigns absent from the workbook (or with empty/invalid
        # tags that did not parse) keep their existing active mapping —
        # including manual entries (import_batch_id IS NULL).
        incoming_campaign_names = [record.campaign_name for record in records]
        if incoming_campaign_names:
            for start in range(0, len(incoming_campaign_names), _BATCH_SIZE):
                chunk = incoming_campaign_names[start:start + _BATCH_SIZE]
                (
                    self.db.table("wbr_pacvue_campaign_map")
                    .update({"active": False})
                    .eq("profile_id", profile_id)
                    .eq("active", True)
                    .in_("campaign_name", chunk)
                    .execute()
                )
        (
            self.db.table("wbr_pacvue_campaign_map")
            .update({"active": True})
            .eq("import_batch_id", batch_id)
            .execute()
        )
        return len(insert_payloads)
