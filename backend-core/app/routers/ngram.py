import os
import tempfile
import itertools
import time
from datetime import date
from typing import Any

from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
from supabase import Client, create_client

import tempfile

from ..auth import require_user, assert_wbr_profile_tool_access
from ..config import settings
from ..usage_logging import usage_logger
from ..services.ngram import (
    read_backview_path,
    build_campaign_items,
    build_workbook,
    NativeNgramWorkbookService,
)
from ..services.ngram.override_capture import persist_ai_override_capture
from openpyxl import load_workbook
import xlsxwriter

router = APIRouter(prefix="/ngram", tags=["ngram"])


@router.get("/healthz")
def health():
    return {"ok": True}


MAX_UPLOAD_MB = int(os.getenv("MAX_UPLOAD_MB", "40"))


def _get_supabase() -> Client:
    if not settings.supabase_url or not settings.supabase_service_role:
        raise HTTPException(status_code=500, detail="Supabase credentials not configured")
    try:
        return create_client(settings.supabase_url, settings.supabase_service_role)
    except Exception as exc:
        raise HTTPException(status_code=500, detail="Failed to initialize Supabase client") from exc


def _get_native_service() -> NativeNgramWorkbookService:
    return NativeNgramWorkbookService(_get_supabase())


class NativeWorkbookRequest(BaseModel):
    profile_id: str = Field(..., min_length=1)
    ad_product: str = Field(..., min_length=1)
    date_from: date
    date_to: date
    respect_legacy_exclusions: bool = True


class CampaignScratchpadPrefill(BaseModel):
    campaign_name: str = Field(..., min_length=1)
    exact: list[str] = Field(default_factory=list)
    mono: list[str] = Field(default_factory=list)
    bi: list[str] = Field(default_factory=list)
    tri: list[str] = Field(default_factory=list)


class CampaignTermReviewPrefill(BaseModel):
    search_term: str = Field(..., min_length=1)
    recommendation: str = Field(..., min_length=1)
    confidence: str = Field(..., min_length=1)
    reason_tag: str = Field(..., min_length=1)
    rationale: str | None = None


class NativePrefilledWorkbookRequest(NativeWorkbookRequest):
    preview_run_id: str | None = None
    campaign_prefills: list[CampaignScratchpadPrefill] = Field(default_factory=list)
    campaign_term_reviews: dict[str, list[CampaignTermReviewPrefill]] = Field(default_factory=dict)


def _to_non_empty_text(value: Any) -> str:
    return str(value or "").strip()


def _normalize_search_term_key(value: Any) -> str:
    return _to_non_empty_text(value).casefold()


def _build_prefill_context_from_request(
    request: NativePrefilledWorkbookRequest,
) -> tuple[dict[str, dict[str, list[str]]], dict[str, dict[str, dict[str, str | None]]], dict[str, str | float | None] | None]:
    ai_prefills = {
        item.campaign_name: {
            "exact": item.exact,
            "mono": item.mono,
            "bi": item.bi,
            "tri": item.tri,
        }
        for item in request.campaign_prefills
    }
    ai_term_reviews = {
        campaign_name: {
            _normalize_search_term_key(review.search_term): {
                "recommendation": review.recommendation,
                "confidence": review.confidence,
                "reason_tag": review.reason_tag,
                "rationale": review.rationale,
            }
            for review in reviews
            if _normalize_search_term_key(review.search_term)
        }
        for campaign_name, reviews in request.campaign_term_reviews.items()
        if _to_non_empty_text(campaign_name)
    }
    return ai_prefills, ai_term_reviews, None


def _build_prefill_context_from_saved_preview(
    request: NativePrefilledWorkbookRequest,
    preview_run: dict[str, Any],
) -> tuple[dict[str, dict[str, list[str]]], dict[str, dict[str, dict[str, str | None]]], dict[str, str | float | None]]:
    if str(preview_run.get("profile_id") or "") != request.profile_id:
        raise HTTPException(status_code=400, detail="Saved AI preview run does not match the selected profile.")
    if _to_non_empty_text(preview_run.get("ad_product")).upper() != request.ad_product.strip().upper():
        raise HTTPException(status_code=400, detail="Saved AI preview run does not match the selected ad product.")
    if _to_non_empty_text(preview_run.get("date_from")) != request.date_from.isoformat():
        raise HTTPException(status_code=400, detail="Saved AI preview run does not match the selected start date.")
    if _to_non_empty_text(preview_run.get("date_to")) != request.date_to.isoformat():
        raise HTTPException(status_code=400, detail="Saved AI preview run does not match the selected end date.")
    if bool(preview_run.get("respect_legacy_exclusions")) != request.respect_legacy_exclusions:
        raise HTTPException(
            status_code=400,
            detail="Saved AI preview run does not match the selected legacy-exclusion setting.",
        )

    preview_payload = preview_run.get("preview_payload")
    if not isinstance(preview_payload, dict):
        raise HTTPException(status_code=400, detail="Saved AI preview run is missing a valid preview payload.")

    campaigns = preview_payload.get("campaigns")
    if not isinstance(campaigns, list):
        raise HTTPException(status_code=400, detail="Saved AI preview run payload is missing campaigns.")

    ai_prefills: dict[str, dict[str, list[str]]] = {}
    ai_term_reviews: dict[str, dict[str, dict[str, str | None]]] = {}

    for campaign in campaigns:
        if not isinstance(campaign, dict):
            continue
        campaign_name = _to_non_empty_text(campaign.get("campaignName"))
        if not campaign_name:
            continue

        synthesized_prefills = campaign.get("synthesizedPrefills")
        model_prefills = campaign.get("modelPrefills")
        exact_prefills = (
            [
                _to_non_empty_text(value)
                for value in model_prefills.get("exact", [])
                if _to_non_empty_text(value)
            ]
            if isinstance(model_prefills, dict)
            else []
        )
        if isinstance(synthesized_prefills, dict):
            ai_prefills[campaign_name] = {
                "exact": exact_prefills,
                "mono": [
                    _to_non_empty_text(item.get("gram"))
                    for item in synthesized_prefills.get("mono", [])
                    if isinstance(item, dict) and _to_non_empty_text(item.get("gram"))
                ],
                "bi": [
                    _to_non_empty_text(item.get("gram"))
                    for item in synthesized_prefills.get("bi", [])
                    if isinstance(item, dict) and _to_non_empty_text(item.get("gram"))
                ],
                "tri": [
                    _to_non_empty_text(item.get("gram"))
                    for item in synthesized_prefills.get("tri", [])
                    if isinstance(item, dict) and _to_non_empty_text(item.get("gram"))
                ],
            }
        elif isinstance(model_prefills, dict):
            ai_prefills[campaign_name] = {
                "exact": exact_prefills,
                "mono": [
                    _to_non_empty_text(value)
                    for value in model_prefills.get("mono", [])
                    if _to_non_empty_text(value)
                ],
                "bi": [
                    _to_non_empty_text(value)
                    for value in model_prefills.get("bi", [])
                    if _to_non_empty_text(value)
                ],
                "tri": [
                    _to_non_empty_text(value)
                    for value in model_prefills.get("tri", [])
                    if _to_non_empty_text(value)
                ],
            }

        evaluations = campaign.get("evaluations")
        if isinstance(evaluations, list):
            review_lookup: dict[str, dict[str, str | None]] = {}
            for evaluation in evaluations:
                if not isinstance(evaluation, dict):
                    continue
                search_term_key = _normalize_search_term_key(evaluation.get("search_term"))
                if not search_term_key:
                    continue
                review_lookup[search_term_key] = {
                    "recommendation": _to_non_empty_text(evaluation.get("recommendation")),
                    "confidence": _to_non_empty_text(evaluation.get("confidence")),
                    "reason_tag": _to_non_empty_text(evaluation.get("reason_tag")),
                    "rationale": _to_non_empty_text(evaluation.get("rationale")) or None,
                }
            if review_lookup:
                ai_term_reviews[campaign_name] = review_lookup

    ai_summary: dict[str, str | float | None] = {
        "preview_run_id": _to_non_empty_text(preview_run.get("id")) or None,
        "model": _to_non_empty_text(preview_payload.get("model")) or _to_non_empty_text(preview_run.get("model")) or None,
        "prompt_version": _to_non_empty_text(preview_payload.get("prompt_version")) or _to_non_empty_text(preview_run.get("prompt_version")) or None,
        "spend_threshold": float(preview_run.get("spend_threshold")) if preview_run.get("spend_threshold") is not None else None,
        "output_mode": "triage_only",
    }

    return ai_prefills, ai_term_reviews, ai_summary


def _load_saved_preview_run(service: NativeNgramWorkbookService, preview_run_id: str) -> dict[str, Any]:
    response = (
        service.db.table("ngram_ai_preview_runs")
        .select("id,profile_id,ad_product,date_from,date_to,spend_threshold,respect_legacy_exclusions,model,prompt_version,preview_payload,requested_by_auth_user_id")
        .eq("id", preview_run_id)
        .limit(1)
        .execute()
    )
    rows = response.data if isinstance(response.data, list) else []
    if not rows:
        raise HTTPException(status_code=400, detail="Saved AI preview run was not found.")
    if not isinstance(rows[0], dict):
        raise HTTPException(status_code=400, detail="Saved AI preview run payload is invalid.")
    return rows[0]


@router.post("/process", response_class=FileResponse)
async def process_report(
    file: UploadFile = File(...),
    user=Depends(require_user),
):
    started = time.time()
    total = 0
    chunk_size = 2 * 1024 * 1024

    with tempfile.NamedTemporaryFile(delete=False) as tmp:
        tmp_path = tmp.name
        while True:
            chunk = await file.read(chunk_size)
            if not chunk:
                break
            total += len(chunk)
            if total > MAX_UPLOAD_MB * 1024 * 1024:
                await file.close()
                os.unlink(tmp_path)
                raise HTTPException(status_code=413, detail="File too large")
            tmp.write(chunk)
    await file.close()

    try:
        df = read_backview_path(tmp_path, file.filename)
    except Exception as exc:
        os.unlink(tmp_path)
        raise HTTPException(status_code=400, detail=f"Parse error: {exc}") from exc

    os.unlink(tmp_path)

    build_result = build_campaign_items(df, respect_legacy_exclusions=True)
    campaign_items = build_result.campaign_items

    if not campaign_items:
        raise HTTPException(status_code=400, detail="No eligible campaigns after filters (Ex./SD*).")

    workbook_path = build_workbook(campaign_items, settings.app_version)
    dl_name = (
        os.path.splitext(file.filename)[0].replace(" ", "_")
        + "_ngrams.xlsx"
    )

    usage_logger.log(
        {
            "user_id": user.get("sub"),
            "user_email": user.get("email"),
            "tool": "ngram",
            "action": "process",
            "file_name": file.filename,
            "file_size_bytes": total,
            "rows_processed": int(df.shape[0]),
            "campaigns": len(campaign_items),
            "status": "success",
            "duration_ms": int((time.time() - started) * 1000),
            "app_version": settings.app_version,
        }
    )

    try:
        override_capture = persist_ai_override_capture(
            _get_supabase(),
            wb,
            collected_by_auth_user_id=_to_non_empty_text(user.get("sub")) or None,
            source_filename=file.filename,
        )
        if override_capture:
            usage_logger.log(
                {
                    "user_id": user.get("sub"),
                    "user_email": user.get("email"),
                    "tool": "ngram",
                    "action": "collect_ai_override_capture",
                    "file_name": file.filename,
                    "preview_run_id": override_capture.get("preview_run_id"),
                    "override_capture_id": override_capture.get("id"),
                    "status": "success",
                    "app_version": settings.app_version,
                }
            )
    except Exception:
        # Override capture is best-effort and should never block summary export.
        pass

    return FileResponse(
        workbook_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=dl_name,
    )


@router.post("/native-workbook", response_class=FileResponse)
async def build_native_workbook(
    request: NativeWorkbookRequest,
    service: NativeNgramWorkbookService = Depends(_get_native_service),
    user=Depends(require_user),
):
    assert_wbr_profile_tool_access(user, request.profile_id, "ngram-2")

    if request.date_from > request.date_to:
        raise HTTPException(status_code=400, detail="date_from must be on or before date_to")

    started = time.time()

    try:
        result = service.build_workbook_from_search_term_facts(
            profile_id=request.profile_id,
            ad_product=request.ad_product,
            date_from=request.date_from,
            date_to=request.date_to,
            respect_legacy_exclusions=request.respect_legacy_exclusions,
            app_version=settings.app_version,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail="Failed to generate native N-Gram workbook") from exc

    usage_logger.log(
        {
            "user_id": user.get("sub"),
            "user_email": user.get("email"),
            "tool": "ngram",
            "action": "native_workbook",
            "profile_id": request.profile_id,
            "ad_product": result.ad_product,
            "date_from": request.date_from.isoformat(),
            "date_to": request.date_to.isoformat(),
            "rows_processed": result.rows_processed,
            "campaigns": result.campaigns_included,
            "campaigns_skipped": result.campaigns_skipped,
            "status": "success",
            "duration_ms": int((time.time() - started) * 1000),
            "app_version": settings.app_version,
        }
    )

    return FileResponse(
        result.workbook_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=result.filename,
    )


@router.post("/native-workbook-prefilled", response_class=FileResponse)
async def build_native_prefilled_workbook(
    request: NativePrefilledWorkbookRequest,
    service: NativeNgramWorkbookService = Depends(_get_native_service),
    user=Depends(require_user),
):
    assert_wbr_profile_tool_access(user, request.profile_id, "ngram-2")

    if request.date_from > request.date_to:
        raise HTTPException(status_code=400, detail="date_from must be on or before date_to")

    started = time.time()

    if request.preview_run_id:
        saved_preview_run = _load_saved_preview_run(service, request.preview_run_id)
        requested_by_auth_user_id = _to_non_empty_text(saved_preview_run.get("requested_by_auth_user_id"))
        if requested_by_auth_user_id and requested_by_auth_user_id != _to_non_empty_text(user.get("sub")):
            raise HTTPException(status_code=403, detail="Saved AI preview run belongs to a different user.")
        ai_prefills, ai_term_reviews, ai_summary = _build_prefill_context_from_saved_preview(
            request,
            saved_preview_run,
        )
    else:
        ai_prefills, ai_term_reviews, ai_summary = _build_prefill_context_from_request(request)

    try:
        result = service.build_workbook_from_search_term_facts(
            profile_id=request.profile_id,
            ad_product=request.ad_product,
            date_from=request.date_from,
            date_to=request.date_to,
            respect_legacy_exclusions=request.respect_legacy_exclusions,
            app_version=settings.app_version,
            ai_prefills=ai_prefills,
            ai_term_reviews=ai_term_reviews,
            ai_summary=ai_summary,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail="Failed to generate AI-prefilled N-Gram workbook") from exc

    usage_logger.log(
        {
            "user_id": user.get("sub"),
            "user_email": user.get("email"),
            "tool": "ngram",
            "action": "native_prefilled_workbook",
            "profile_id": request.profile_id,
            "ad_product": result.ad_product,
            "date_from": request.date_from.isoformat(),
            "date_to": request.date_to.isoformat(),
            "rows_processed": result.rows_processed,
            "campaigns": result.campaigns_included,
            "campaigns_skipped": result.campaigns_skipped,
            "prefill_campaigns": len(ai_prefills),
            "prefill_mono": sum(len(item.get("mono", [])) for item in ai_prefills.values()),
            "prefill_bi": sum(len(item.get("bi", [])) for item in ai_prefills.values()),
            "prefill_tri": sum(len(item.get("tri", [])) for item in ai_prefills.values()),
            "preview_run_id": request.preview_run_id,
            "status": "success",
            "duration_ms": int((time.time() - started) * 1000),
            "app_version": settings.app_version,
        }
    )

    return FileResponse(
        result.workbook_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=result.filename,
    )


@router.post("/native-summary")
async def build_native_summary(
    request: NativeWorkbookRequest,
    service: NativeNgramWorkbookService = Depends(_get_native_service),
    user=Depends(require_user),
):
    assert_wbr_profile_tool_access(user, request.profile_id, "ngram-2")

    if request.date_from > request.date_to:
        raise HTTPException(status_code=400, detail="date_from must be on or before date_to")

    started = time.time()

    try:
        result = service.build_summary_from_search_term_facts(
            profile_id=request.profile_id,
            ad_product=request.ad_product,
            date_from=request.date_from,
            date_to=request.date_to,
            respect_legacy_exclusions=request.respect_legacy_exclusions,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail="Failed to build native N-Gram summary") from exc

    usage_logger.log(
        {
            "user_id": user.get("sub"),
            "user_email": user.get("email"),
            "tool": "ngram",
            "action": "native_summary",
            "profile_id": request.profile_id,
            "ad_product": result.ad_product,
            "date_from": request.date_from.isoformat(),
            "date_to": request.date_to.isoformat(),
            "rows_processed": result.eligible_rows,
            "campaigns": result.campaigns_included,
            "campaigns_skipped": result.campaigns_skipped,
            "status": "success",
            "duration_ms": int((time.time() - started) * 1000),
            "app_version": settings.app_version,
        }
    )

    return {
        "ok": True,
        "summary": {
            "ad_product": result.ad_product,
            "profile_id": result.profile_id,
            "profile_display_name": result.profile_display_name,
            "marketplace_code": result.marketplace_code,
            "date_from": result.date_from,
            "date_to": result.date_to,
            "raw_rows": result.raw_rows,
            "eligible_rows": result.eligible_rows,
            "excluded_asin_rows": result.excluded_asin_rows,
            "excluded_incomplete_rows": result.excluded_incomplete_rows,
            "unique_campaigns": result.unique_campaigns,
            "unique_search_terms": result.unique_search_terms,
            "campaigns_included": result.campaigns_included,
            "campaigns_skipped": result.campaigns_skipped,
            "report_dates_present": result.report_dates_present,
            "coverage_start": result.coverage_start,
            "coverage_end": result.coverage_end,
            "imported_totals": {
                "impressions": result.imported_totals.impressions,
                "clicks": result.imported_totals.clicks,
                "spend": result.imported_totals.spend,
                "orders": result.imported_totals.orders,
                "sales": result.imported_totals.sales,
            },
            "workbook_input_totals": {
                "impressions": result.workbook_input_totals.impressions,
                "clicks": result.workbook_input_totals.clicks,
                "spend": result.workbook_input_totals.spend,
                "orders": result.workbook_input_totals.orders,
                "sales": result.workbook_input_totals.sales,
            },
            "campaigns": [
                {
                    "campaign_name": campaign.campaign_name,
                    "search_terms": campaign.search_terms,
                    "spend": campaign.spend,
                }
                for campaign in result.campaigns
            ],
            "warnings": result.warnings,
        },
    }


@router.post("/collect", response_class=FileResponse)
async def collect_negatives(
    file: UploadFile = File(...),
    user=Depends(require_user),
):
    # Parse a filled workbook and extract NE keywords plus scratchpad mono/bi/tri
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
        tmp.write(await file.read())
    await file.close()

    try:
        wb = load_workbook(tmp_path, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Unable to read workbook: {exc}") from exc
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass

    per_campaign: dict[str, dict[str, list[str]]] = {}

    def _append_unique(values: list[str], candidate: object) -> None:
        text = str(candidate or "").strip()
        if not text or text in values:
            return
        values.append(text)

    def _last_non_empty(sheet, col: str, start_row: int) -> int:
        max_row = sheet.max_row
        for i in range(max_row, start_row - 1, -1):
            val = sheet[f"{col}{i}"].value
            if val not in (None, ""):
                return i
        return start_row - 1

    for sheet in wb.worksheets:
        if sheet.title in {"Summary", "NE Summary"}:
            continue
        campaign_name = sheet["B1"].value or sheet.title
        bucket = per_campaign.setdefault(
            str(campaign_name),
            {"ne": [], "mono": [], "bi": [], "tri": []},
        )
        last_ne_row = _last_non_empty(sheet, "AT", 2)
        last_mono_row = _last_non_empty(sheet, "K", 7)  # Monogram NE/NP column
        last_bi_row = _last_non_empty(sheet, "X", 7)   # Bigram NE/NP column
        last_tri_row = _last_non_empty(sheet, "AK", 7)  # Trigram NE/NP column
        last_scratch_mono_row = _last_non_empty(sheet, "AZ", 7)
        last_scratch_bi_row = _last_non_empty(sheet, "BA", 7)
        last_scratch_tri_row = _last_non_empty(sheet, "BB", 7)
        # NE keywords from search term table (AN with AT = "NE")
        for i in range(2, last_ne_row + 1):
            flag = sheet[f"AT{i}"].value
            term = sheet[f"AN{i}"].value
            if (flag or "").strip().upper() == "NE" and term not in (None, ""):
                _append_unique(bucket["ne"], term)
        # Monogram/Bigram/Trigram NE/NP column flags.
        # These can remain blank if the workbook is uploaded before Excel
        # recalculates formulas, so scratchpad columns are also read below.
        for i in range(7, last_mono_row + 1):
            flag = sheet[f"K{i}"].value
            gram = sheet[f"A{i}"].value
            if (flag or "").strip().upper() in {"NE", "NP"} and gram not in (None, ""):
                _append_unique(bucket["mono"], gram)
        for i in range(7, last_bi_row + 1):
            flag = sheet[f"X{i}"].value
            gram = sheet[f"N{i}"].value
            if (flag or "").strip().upper() in {"NE", "NP"} and gram not in (None, ""):
                _append_unique(bucket["bi"], gram)
        for i in range(7, last_tri_row + 1):
            flag = sheet[f"AK{i}"].value
            gram = sheet[f"AA{i}"].value
            if (flag or "").strip().upper() in {"NE", "NP"} and gram not in (None, ""):
                _append_unique(bucket["tri"], gram)
        # Scratchpad source columns are the durable source of truth for AI-
        # prefilled or analyst-entered grams and do not depend on formula
        # recalculation.
        for i in range(7, last_scratch_mono_row + 1):
            _append_unique(bucket["mono"], sheet[f"AZ{i}"].value)
        for i in range(7, last_scratch_bi_row + 1):
            _append_unique(bucket["bi"], sheet[f"BA{i}"].value)
        for i in range(7, last_scratch_tri_row + 1):
            _append_unique(bucket["tri"], sheet[f"BB{i}"].value)

    rows_out = [["Campaign", "NE Keywords", "Monogram", "Bigram", "Trigram"]]
    for campaign_name, data in per_campaign.items():
        if not (data["ne"] or data["mono"] or data["bi"] or data["tri"]):
            continue
        for ne, mono, bi, tri in itertools.zip_longest(
            data["ne"], data["mono"], data["bi"], data["tri"], fillvalue=""
        ):
            rows_out.append([campaign_name, ne, mono, bi, tri])

    if len(rows_out) == 1:
        raise HTTPException(status_code=400, detail="No NE or scratchpad entries found.")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as out_tmp:
        out_path = out_tmp.name

    workbook = xlsxwriter.Workbook(out_path)
    ws = workbook.add_worksheet("NE Summary")
    header_fmt = workbook.add_format(
        {"bold": True, "bg_color": "#0066CC", "font_color": "#FFFFFF", "border": 1, "align": "center", "valign": "vcenter"}
    )
    border_fmt = workbook.add_format({"border": 1})
    zebra_fmt = workbook.add_format({"bg_color": "#F2F2F2", "border": 1})

    for j, col_name in enumerate(rows_out[0]):
        ws.write_string(0, j, col_name, header_fmt)

    for r, row_vals in enumerate(rows_out[1:], start=1):
        fmt = zebra_fmt if r % 2 == 0 else border_fmt
        for c, val in enumerate(row_vals):
            ws.write(r, c, val, fmt)

    ws.set_column("A:A", 50)
    ws.set_column("B:B", 60)
    ws.set_column("C:E", 30)
    ws.freeze_panes(1, 0)
    workbook.close()

    usage_logger.log(
        {
            "user_id": user.get("sub"),
            "user_email": user.get("email"),
            "tool": "ngram",
            "action": "collect",
            "file_name": file.filename,
            "status": "success",
            "rows_emitted": len(rows_out) - 1,
            "app_version": settings.app_version,
        }
    )

    try:
        override_capture = persist_ai_override_capture(
            _get_supabase(),
            wb,
            collected_by_auth_user_id=_to_non_empty_text(user.get("sub")) or None,
            source_filename=file.filename,
        )
        if override_capture:
            usage_logger.log(
                {
                    "user_id": user.get("sub"),
                    "user_email": user.get("email"),
                    "tool": "ngram",
                    "action": "collect_ai_override_capture",
                    "file_name": file.filename,
                    "preview_run_id": override_capture.get("preview_run_id"),
                    "override_capture_id": override_capture.get("id"),
                    "status": "success",
                    "app_version": settings.app_version,
                }
            )
    except Exception as exc:
        usage_logger.log(
            {
                "user_id": user.get("sub"),
                "user_email": user.get("email"),
                "tool": "ngram",
                "action": "collect_ai_override_capture",
                "file_name": file.filename,
                "status": "error",
                "error": str(exc),
                "app_version": settings.app_version,
            }
        )

    return FileResponse(
        out_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"{os.path.splitext(file.filename)[0]}_negatives.xlsx",
    )
