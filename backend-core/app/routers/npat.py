import os
import tempfile
import itertools
import time

from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import FileResponse

from ..auth import require_user
from ..config import settings
from ..usage_logging import usage_logger
from ..services.npat import (
    read_backview_path,
    calculate_asin_metrics,
    derive_category,
    build_npat_workbook,
)
from openpyxl import load_workbook
import xlsxwriter

router = APIRouter(prefix="/npat", tags=["npat"])


@router.get("/healthz")
def health():
    """Health check endpoint."""
    return {"ok": True}


MAX_UPLOAD_MB = int(os.getenv("MAX_UPLOAD_MB", "40"))


@router.post("/process", response_class=FileResponse)
async def process_report(
    file: UploadFile = File(...),
    user=Depends(require_user),
):
    """
    Generate N-PAT workbook from Search Term Report.

    Filters to ONLY ASINs, groups by campaign, calculates metrics,
    and generates Excel with H10 integration zones.
    """
    started = time.time()
    total = 0
    chunk_size = 2 * 1024 * 1024

    with tempfile.NamedTemporaryFile(delete=False) as tmp:
        tmp_path = tmp.name
        while True:
            chunk = await file.read(chunk_size)
            if not chunk:
                break
            total += len(chunk)
            if total > MAX_UPLOAD_MB * 1024 * 1024:
                await file.close()
                os.unlink(tmp_path)
                raise HTTPException(status_code=413, detail="File too large")
            tmp.write(chunk)
    await file.close()

    try:
        df = read_backview_path(tmp_path, file.filename)
    except Exception as exc:
        os.unlink(tmp_path)
        raise HTTPException(status_code=400, detail=f"Parse error: {exc}") from exc

    os.unlink(tmp_path)

    if df.empty:
        raise HTTPException(
            status_code=400,
            detail="No ASINs found in uploaded file. N-PAT requires ASIN data (10-character alphanumeric codes like B08XYZ123)."
        )

    campaign_items = []
    for camp, sub in df.groupby("Campaign Name"):
        cname = str(camp)
        # Exclude campaigns with "Ex.", "SDI", or "SDV" in name
        if any(x in cname for x in ["Ex.", "SDI", "SDV"]):
            continue

        category_raw, category_key, cat_notes = derive_category(cname)

        # Calculate ASIN metrics
        asins = calculate_asin_metrics(sub)

        notes = list(cat_notes)
        if asins.empty:
            notes.append("No ASINs found for this campaign.")

        campaign_items.append(
            {
                "campaign_name": cname,
                "category_raw": category_raw,
                "category_key": category_key,
                "asins": asins,
                "notes": notes,
            }
        )

    if not campaign_items:
        raise HTTPException(status_code=400, detail="No eligible campaigns after filters (Ex./SD*).")

    workbook_path = build_npat_workbook(campaign_items, settings.app_version)
    dl_name = (
        os.path.splitext(file.filename)[0].replace(" ", "_")
        + "_npat.xlsx"
    )

    # Calculate total ASINs across all campaigns
    total_asins = sum(len(item["asins"]) for item in campaign_items)

    usage_logger.log(
        {
            "user_id": user.get("sub"),
            "user_email": user.get("email"),
            "file_name": file.filename,
            "file_size_bytes": total,
            "rows_processed": int(df.shape[0]),
            "campaigns": len(campaign_items),
            "total_asins": total_asins,
            "status": "success",
            "duration_ms": int((time.time() - started) * 1000),
            "app_version": settings.app_version,
        }
    )

    return FileResponse(
        workbook_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=dl_name,
    )


@router.post("/collect", response_class=FileResponse)
async def collect_negatives(
    file: UploadFile = File(...),
    user=Depends(require_user),
):
    """
    Extract negatives summary from filled N-PAT workbook.

    Reads ASINs marked "NE" in column Q and generates simple
    2-column summary (Campaign, NE ASIN) matching N-Gram pattern.
    """
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
        tmp.write(await file.read())
    await file.close()

    try:
        wb = load_workbook(tmp_path, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Unable to read workbook: {exc}") from exc
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass

    # Per-campaign negative ASINs (simple list)
    per_campaign: dict[str, list[str]] = {}

    def _last_non_empty(sheet, col: str, start_row: int) -> int:
        """Find last non-empty row in column."""
        max_row = sheet.max_row
        for i in range(max_row, start_row - 1, -1):
            val = sheet[f"{col}{i}"].value
            if val not in (None, ""):
                return i
        return start_row - 1

    for sheet in wb.worksheets:
        if sheet.title in {"Summary"}:
            continue

        # Get campaign name from cell B1
        campaign_name = sheet["B1"].value or sheet.title

        # Find last row with data in NE/NP column (Q) - data starts at row 7
        last_ne_row = _last_non_empty(sheet, "Q", 7)

        negatives = []
        # Read ASINs marked "NE" in column Q (starting from row 7)
        for i in range(7, last_ne_row + 1):
            flag = sheet[f"Q{i}"].value
            if (flag or "").strip().upper() == "NE":
                asin = sheet[f"A{i}"].value
                if asin in (None, ""):
                    continue
                negatives.append(str(asin))

        if negatives:
            per_campaign[str(campaign_name)] = negatives

    if not per_campaign:
        raise HTTPException(status_code=400, detail="No NE markings found. Please mark ASINs as 'NE' in column Q before uploading.")

    # Generate negatives summary Excel (simple 2-column format matching N-Gram)
    rows_out = [["Campaign", "NE ASIN"]]

    for campaign_name, asins in per_campaign.items():
        for asin in asins:
            rows_out.append([campaign_name, asin])

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as out_tmp:
        out_path = out_tmp.name

    workbook = xlsxwriter.Workbook(out_path)
    ws = workbook.add_worksheet("NE Summary")

    # Define formats (simple, matching N-Gram)
    header_fmt = workbook.add_format(
        {"bold": True, "bg_color": "#0066CC", "font_color": "#FFFFFF", "border": 1, "align": "center", "valign": "vcenter"}
    )
    border_fmt = workbook.add_format({"border": 1})
    zebra_fmt = workbook.add_format({"bg_color": "#F2F2F2", "border": 1})

    # Write headers
    for j, col_name in enumerate(rows_out[0]):
        ws.write_string(0, j, col_name, header_fmt)

    # Write data rows (simple text format)
    for r, row_vals in enumerate(rows_out[1:], start=1):
        fmt = zebra_fmt if r % 2 == 0 else border_fmt
        for c, val in enumerate(row_vals):
            ws.write(r, c, val, fmt)

    # Set column widths
    ws.set_column("A:A", 50)  # Campaign
    ws.set_column("B:B", 15)  # NE ASIN

    ws.freeze_panes(1, 0)
    workbook.close()

    usage_logger.log(
        {
            "user_id": user.get("sub"),
            "user_email": user.get("email"),
            "file_name": file.filename,
            "status": "success",
            "rows_emitted": len(rows_out) - 1,
            "app_version": settings.app_version,
        }
    )

    return FileResponse(
        out_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"{os.path.splitext(file.filename)[0]}_negatives.xlsx",
    )
