"""Tests for WBR Pacvue import parsing and service behavior."""

from __future__ import annotations

import io
import zipfile
from unittest.mock import MagicMock

import pytest
from openpyxl import Workbook

from app.services.wbr.pacvue_imports import PacvueImportService, parse_pacvue_workbook
from app.services.wbr.profiles import WBRValidationError


def _build_workbook_bytes(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_multisheet_workbook_bytes(sheet_rows: list[tuple[str, list[list[object]]]]) -> bytes:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for title, rows in sheet_rows:
        sheet = workbook.create_sheet(title)
        for row in rows:
            sheet.append(row)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _force_bad_dimension(file_bytes: bytes, sheet_name: str = "xl/worksheets/sheet1.xml") -> bytes:
    source = io.BytesIO(file_bytes)
    output = io.BytesIO()

    with zipfile.ZipFile(source, "r") as zin, zipfile.ZipFile(output, "w") as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == sheet_name:
                data = data.replace(b'<dimension ref="A1:C3"/>', b'<dimension ref="A1"/>')
                data = data.replace(b'<dimension ref="A1:C4"/>', b'<dimension ref="A1"/>')
                data = data.replace(b'<dimension ref="A1:C6"/>', b'<dimension ref="A1"/>')
            zout.writestr(item, data)

    return output.getvalue()


def _chain_table(response_data: list[dict] | None = None) -> MagicMock:
    table = MagicMock()
    table.select.return_value = table
    table.insert.return_value = table
    table.update.return_value = table
    table.delete.return_value = table
    table.eq.return_value = table
    table.order.return_value = table
    table.limit.return_value = table
    resp = MagicMock()
    resp.data = response_data if response_data is not None else []
    table.execute.return_value = resp
    return table


def _multi_table_db(mapping: dict[str, list[MagicMock]]) -> MagicMock:
    iterators = {name: iter(tables) for name, tables in mapping.items()}

    def router(name: str) -> MagicMock:
        return next(iterators[name])

    db = MagicMock()
    db.table.side_effect = router
    return db


class TestParsePacvueWorkbook:
    def test_detects_header_after_metadata_rows(self):
        file_bytes = _build_workbook_bytes(
            [
                ["Level", "Campaign"],
                ["Time Range", "03/02/2026 - 03/08/2026"],
                ["Download Time", "2026-03-12 12:21:38"],
                [],
                ["Name", "state", "CampaignTagNames"],
                ["Screen Shine - Duo | SPM", "enabled", "Screen Shine | Duo / Perf"],
            ]
        )

        parsed = parse_pacvue_workbook(file_bytes)

        assert parsed.header_row_index == 4
        assert parsed.rows_read == 1
        assert parsed.records[0].campaign_name == "Screen Shine - Duo | SPM"
        assert parsed.records[0].leaf_row_label == "Screen Shine | Duo"
        assert parsed.records[0].goal_code == "Perf"

    def test_dedupes_identical_campaign_rows(self):
        file_bytes = _build_workbook_bytes(
            [
                ["Name", "CampaignTagNames"],
                ["Campaign A", "Screen Shine | Pro / Perf"],
                ["Campaign A", "Screen Shine | Pro / Perf"],
            ]
        )

        parsed = parse_pacvue_workbook(file_bytes)

        assert parsed.rows_read == 2
        assert parsed.duplicate_rows_skipped == 1
        assert len(parsed.records) == 1

    def test_detects_header_on_non_active_sheet(self):
        file_bytes = _build_multisheet_workbook_bytes(
            [
                ("Intro", [["Level", "Campaign"], ["Download Time", "2026-03-12 12:21:38"]]),
                (
                    "Campaigns",
                    [
                        ["Level", "Campaign"],
                        ["Time Range", "03/02/2026 - 03/08/2026"],
                        [],
                        ["Name", "state", "CampaignTagNames"],
                        ["Screen Shine - Duo | SPM", "enabled", "Screen Shine | Duo / Perf"],
                    ],
                ),
            ]
        )

        parsed = parse_pacvue_workbook(file_bytes)

        assert parsed.sheet_title == "Campaigns"
        assert parsed.header_row_index == 3
        assert parsed.records[0].campaign_name == "Screen Shine - Duo | SPM"

    def test_rejects_conflicting_duplicate_campaign_rows(self):
        file_bytes = _build_workbook_bytes(
            [
                ["Name", "CampaignTagNames"],
                ["Campaign A", "Screen Shine | Pro / Perf"],
                ["Campaign A", "Screen Shine | Pro / Harv"],
            ]
        )

        with pytest.raises(WBRValidationError, match="conflicting Pacvue tags"):
            parse_pacvue_workbook(file_bytes)

    def test_rejects_unsupported_goal_suffix(self):
        file_bytes = _build_workbook_bytes(
            [
                ["Name", "CampaignTagNames"],
                ["Campaign A", "Screen Shine | Pro / Weird"],
            ]
        )

        with pytest.raises(WBRValidationError, match="unsupported goal suffix"):
            parse_pacvue_workbook(file_bytes)

    def test_accepts_category_as_comp_goal_suffix(self):
        file_bytes = _build_workbook_bytes(
            [
                ["Name", "CampaignTagNames"],
                ["Campaign A", "Screen Shine | Pro / Category"],
            ]
        )

        parsed = parse_pacvue_workbook(file_bytes)

        assert parsed.records[0].goal_code == "Comp"

    def test_handles_bad_sheet_dimension_metadata(self):
        file_bytes = _build_workbook_bytes(
            [
                ["Level", "Campaign", None],
                ["Time Range", "03/02/2026 - 03/08/2026", None],
                [],
                ["Name", "state", "CampaignTagNames"],
                ["Screen Shine - Duo | SPM", "enabled", "Screen Shine | Duo / Perf"],
            ]
        )
        broken_bytes = _force_bad_dimension(file_bytes)

        parsed = parse_pacvue_workbook(broken_bytes)

        assert parsed.header_row_index == 3
        assert parsed.rows_read == 1
        assert parsed.records[0].leaf_row_label == "Screen Shine | Duo"

    def test_skips_placeholder_unmapped_tags(self):
        file_bytes = _build_workbook_bytes(
            [
                ["Name", "state", "CampaignTagNames"],
                ["Mapped Campaign", "enabled", "Screen Shine | Duo / Perf"],
                ["Unmapped Campaign", "paused", "--"],
            ]
        )

        parsed = parse_pacvue_workbook(file_bytes)

        assert parsed.rows_read == 1
        assert parsed.unmapped_rows_skipped == 1
        assert len(parsed.records) == 1
        assert parsed.records[0].campaign_name == "Mapped Campaign"

    def test_skips_total_footer_row_without_tag(self):
        file_bytes = _build_workbook_bytes(
            [
                ["Name", "state", "CampaignTagNames"],
                ["Mapped Campaign", "enabled", "Screen Shine | Duo / Perf"],
                ["total:", None, None],
            ]
        )

        parsed = parse_pacvue_workbook(file_bytes)

        assert parsed.rows_read == 1
        assert len(parsed.records) == 1
        assert parsed.records[0].campaign_name == "Mapped Campaign"


class TestPacvueImportService:
    def test_import_reactivates_existing_leaf_and_refreshes_mappings(self):
        file_bytes = _build_workbook_bytes(
            [
                ["Name", "CampaignTagNames"],
                ["Campaign A", "Screen Shine | Pro / Perf"],
            ]
        )
        profile = {"id": "p1"}
        batch = {"id": "b1", "import_status": "running"}
        finished_batch = {"id": "b1", "import_status": "success", "rows_read": 1, "rows_loaded": 1}
        inactive_leaf = {"id": "r1", "row_label": "Screen Shine | Pro", "active": False, "sort_order": 10}
        active_leaf = {"id": "r1", "row_label": "Screen Shine | Pro"}

        db = _multi_table_db(
            {
                "wbr_profiles": [_chain_table([profile])],
                "wbr_pacvue_import_batches": [
                    _chain_table([batch]),
                    _chain_table([finished_batch]),
                ],
                "wbr_rows": [
                    _chain_table([inactive_leaf]),   # existing leaf query
                    _chain_table([inactive_leaf]),   # reactivation update
                    _chain_table([active_leaf]),     # active leaf lookup for mapping insert
                ],
                "wbr_pacvue_campaign_map": [
                    _chain_table([{"id": "m1"}]),    # insert inactive mappings
                    _chain_table([]),                # deactivate old active mappings
                    _chain_table([{"id": "m1"}]),    # activate new batch mappings
                ],
            }
        )

        svc = PacvueImportService(db)
        result = svc.import_workbook(
            profile_id="p1",
            file_name="pacvue.xlsx",
            file_bytes=file_bytes,
            user_id="u1",
        )

        assert result["batch"]["import_status"] == "success"
        assert result["summary"]["rows_loaded"] == 1
        assert result["summary"]["reactivated_leaf_rows"] == 1
        assert result["summary"]["created_leaf_rows"] == 0
        assert result["summary"]["unmapped_rows_skipped"] == 0

    def test_rejects_non_xlsx_files(self):
        svc = PacvueImportService(MagicMock())

        with pytest.raises(WBRValidationError, match=".xlsx and .xlsm"):
            svc._validate_file_name("pacvue.csv")
