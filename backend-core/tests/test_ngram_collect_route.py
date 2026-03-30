from __future__ import annotations

import io
import os

import pandas as pd
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app
from app.routers import ngram
from app.services.ngram.workbook import build_workbook


def _make_ngram_df(grams: list[str]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "N-gram": gram,
                "Impression": 10,
                "Click": 2,
                "Spend": 1.5,
                "Order 14d": 0,
                "Sales 14d": 0,
                "CTR": 0.2,
                "CVR": 0.0,
                "CPC": 0.75,
                "ACOS": 0.0,
                "NE/NP": "",
                "Comments": "",
            }
            for gram in grams
        ]
    )


def _override_user():
    return {"sub": "user-123", "email": "tester@example.com"}


def test_collect_route_reads_bigram_scratchpad_without_excel_recalc(monkeypatch):
    campaign_name = "Screen Shine - Duo | SPA | Cls. | Rsrch"
    workbook_path = build_workbook(
        [
            {
                "campaign_name": campaign_name,
                "category_raw": "Duo",
                "category_key": "duo",
                "mono": _make_ngram_df([]),
                "bi": _make_ngram_df(["screen spray", "laptop cloth", "spray bottle"]),
                "tri": _make_ngram_df([]),
                "raw": pd.DataFrame(
                    [
                        {
                            "Search Term": "screen spray",
                            "Impression": 10,
                            "Click": 2,
                            "Spend": 1.5,
                            "Order 14d": 0,
                            "Sales 14d": 0,
                            "NE/NP": "",
                            "Comments": "",
                        }
                    ]
                ),
                "notes": [],
            }
        ],
        "test-version",
        ai_prefills={
            campaign_name: {
                "mono": [],
                "bi": ["screen spray", "laptop cloth", "spray bottle"],
                "tri": [],
            }
        },
        ai_summary={
            "preview_run_id": "preview-run-123",
            "model": "gpt-5.4-2026-03-05",
            "prompt_version": "ngram_step3_calibrated_v2026_03_30",
            "spend_threshold": "1.0",
        },
    )

    monkeypatch.setattr(ngram, "persist_ai_override_capture", lambda *args, **kwargs: None)
    monkeypatch.setattr(ngram, "_get_supabase", lambda: object())
    app.dependency_overrides[ngram.require_user] = _override_user

    try:
        with open(workbook_path, "rb") as handle:
            workbook_bytes = handle.read()

        with TestClient(app) as client:
            response = client.post(
                "/ngram/collect",
                files={
                    "file": (
                        "ai_prefilled.xlsx",
                        io.BytesIO(workbook_bytes),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
    finally:
        app.dependency_overrides.pop(ngram.require_user, None)
        os.unlink(workbook_path)

    assert response.status_code == 200

    summary_wb = load_workbook(io.BytesIO(response.content), data_only=True)
    summary_ws = summary_wb["NE Summary"]
    observed_bigrams = [summary_ws[f"D{row}"].value for row in range(2, summary_ws.max_row + 1)]

    assert observed_bigrams[:3] == ["screen spray", "laptop cloth", "spray bottle"]
