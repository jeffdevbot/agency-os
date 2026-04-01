from __future__ import annotations

import os

import pandas as pd
from openpyxl import load_workbook

from app.services.ngram.override_capture import _build_override_payload, persist_ai_override_capture
from app.services.ngram.workbook import build_workbook


def _make_ngram_df(grams: list[str]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "N-gram": gram,
                "Impression": 10,
                "Click": 2,
                "Spend": 1.5,
                "Order 14d": 0,
                "Sales 14d": 0,
                "CTR": 0.2,
                "CVR": 0.0,
                "CPC": 0.75,
                "ACOS": 0.0,
                "NE/NP": "",
                "Comments": "",
            }
            for gram in grams
        ]
    )


def test_build_override_payload_compares_ai_recommendations_to_reviewed_workbook():
    raw = pd.DataFrame(
        [
            {
                "Search Term": "laptop cloth",
                "Impression": 100,
                "Click": 10,
                "Spend": 8.5,
                "Order 14d": 0,
                "Sales 14d": 0,
                "NE/NP": "",
                "Comments": "",
            },
            {
                "Search Term": "screen cleaner spray",
                "Impression": 80,
                "Click": 8,
                "Spend": 6.2,
                "Order 14d": 0,
                "Sales 14d": 0,
                "NE/NP": "",
                "Comments": "",
            },
        ]
    )

    campaign_name = "Screen Shine - Duo | SPA | Cls. | Rsrch"
    workbook_path = build_workbook(
        [
            {
                "campaign_name": campaign_name,
                "category_raw": "Duo",
                "category_key": "duo",
                "mono": _make_ngram_df(["laptop", "screen"]),
                "bi": _make_ngram_df(["laptop cloth", "screen cleaner"]),
                "tri": _make_ngram_df(["screen cleaner spray"]),
                "raw": raw,
                "notes": [],
            }
        ],
        "test-version",
        ai_term_reviews={
            campaign_name: {
                "laptop cloth": {
                    "recommendation": "NEGATE",
                    "confidence": "HIGH",
                    "reason_tag": "accessory_only_intent",
                },
                "screen cleaner spray": {
                    "recommendation": "KEEP",
                    "confidence": "HIGH",
                    "reason_tag": "core_use_case",
                },
            }
        },
        ai_summary={
            "preview_run_id": "preview-run-123",
            "model": "gpt-5.4-2026-03-05",
            "prompt_version": "ngram_step3_calibrated_v2026_03_30",
            "spend_threshold": "1.0",
        },
    )

    try:
        reviewed_wb = load_workbook(workbook_path, data_only=True)
        payload = _build_override_payload(
            reviewed_wb,
            {
                "id": "preview-run-123",
                "profile_id": "profile-1",
                "model": "gpt-5.4-2026-03-05",
                "prompt_version": "ngram_step3_calibrated_v2026_03_30",
                "preview_payload": {
                    "campaigns": [
                        {
                            "campaignName": campaign_name,
                            "modelPrefills": {
                                "exact": [],
                                "mono": [],
                                "bi": ["laptop cloth"],
                                "tri": [],
                            },
                            "synthesizedPrefills": {
                                "mono": [],
                                "bi": [{"gram": "laptop cloth"}],
                                "tri": [],
                            },
                            "evaluations": [
                                {
                                    "search_term": "laptop cloth",
                                    "recommendation": "NEGATE",
                                    "confidence": "HIGH",
                                    "reason_tag": "accessory_only_intent",
                                },
                                {
                                    "search_term": "screen cleaner spray",
                                    "recommendation": "KEEP",
                                    "confidence": "HIGH",
                                    "reason_tag": "core_use_case",
                                },
                            ],
                        }
                    ]
                },
            },
        )

        assert payload is not None
        assert payload["prompt_version"] == "ngram_step3_calibrated_v2026_03_30"
        assert payload["summary"]["term_status_counts"]["ai_negate_accepted"] == 1
        assert payload["summary"]["term_status_counts"]["unchanged_non_negative"] == 1

        campaign_payload = payload["campaigns"][0]
        laptop_diff = next(item for item in campaign_payload["term_diffs"] if item["search_term"] == "laptop cloth")
        keep_diff = next(
            item for item in campaign_payload["term_diffs"] if item["search_term"] == "screen cleaner spray"
        )

        assert laptop_diff["analyst_outcome"] == "negated"
        assert laptop_diff["analyst_source"] == "gram"
        assert laptop_diff["ai_negation_path"] == "phrase"
        assert laptop_diff["decision_status"] == "matched"
        assert keep_diff["analyst_outcome"] == "not_negated"
        assert keep_diff["decision_status"] == "matched"

        bi_diff = next(item for item in campaign_payload["gram_diffs"]["bi"] if item["gram"] == "laptop cloth")
        assert bi_diff["status"] == "matched"
    finally:
        os.unlink(workbook_path)


def test_build_override_payload_marks_explicit_exact_negatives():
    campaign_name = "Screen Shine - Duo | SPA | Cls. | Rsrch"
    workbook_path = build_workbook(
        [
            {
                "campaign_name": campaign_name,
                "category_raw": "Duo",
                "category_key": "duo",
                "mono": _make_ngram_df([]),
                "bi": _make_ngram_df([]),
                "tri": _make_ngram_df([]),
                "raw": pd.DataFrame(
                    [
                        {
                            "Search Term": "portable monitor travel case",
                            "Impression": 100,
                            "Click": 10,
                            "Spend": 8.5,
                            "Order 14d": 0,
                            "Sales 14d": 0,
                            "NE/NP": "",
                            "Comments": "",
                        }
                    ]
                ),
                "notes": [],
            }
        ],
        "test-version",
        ai_prefills={
            campaign_name: {
                "exact": ["portable monitor travel case"],
                "mono": [],
                "bi": [],
                "tri": [],
            }
        },
        ai_term_reviews={
            campaign_name: {
                "portable monitor travel case": {
                    "recommendation": "NEGATE",
                    "confidence": "HIGH",
                    "reason_tag": "accessory_only_intent",
                }
            }
        },
        ai_summary={
            "preview_run_id": "preview-run-exact",
            "model": "gpt-5.4-2026-03-05",
            "prompt_version": "ngram_pure_model_single_campaign_v2026_03_31",
            "spend_threshold": "1.0",
        },
    )

    try:
        reviewed_wb = load_workbook(workbook_path, data_only=True)
        payload = _build_override_payload(
            reviewed_wb,
            {
                "id": "preview-run-exact",
                "profile_id": "profile-1",
                "model": "gpt-5.4-2026-03-05",
                "prompt_version": "ngram_pure_model_single_campaign_v2026_03_31",
                "preview_payload": {
                    "campaigns": [
                        {
                            "campaignName": campaign_name,
                            "modelPrefills": {
                                "exact": ["portable monitor travel case"],
                                "mono": [],
                                "bi": [],
                                "tri": [],
                            },
                            "evaluations": [
                                {
                                    "search_term": "portable monitor travel case",
                                    "recommendation": "NEGATE",
                                    "confidence": "HIGH",
                                    "reason_tag": "accessory_only_intent",
                                }
                            ],
                        }
                    ]
                },
            },
        )

        assert payload is not None
        campaign_payload = payload["campaigns"][0]
        assert campaign_payload["ai_exact_negatives"] == ["portable monitor travel case"]
        assert campaign_payload["term_diffs"][0]["ai_negation_path"] == "exact"
    finally:
        os.unlink(workbook_path)


class _FakeExecuteResponse:
    def __init__(self, data):
        self.data = data


class _FakePreviewBuilder:
    def __init__(self, preview_row):
        self.preview_row = preview_row

    def select(self, *_args, **_kwargs):
        return self

    def eq(self, *_args, **_kwargs):
        return self

    def limit(self, *_args, **_kwargs):
        return self

    def execute(self):
        return _FakeExecuteResponse([self.preview_row])


class _FakeInsertBuilder:
    def __init__(self, store: list[dict]):
        self.store = store
        self.payload = None

    def insert(self, payload):
        self.payload = payload
        return self

    def execute(self):
        self.store.append(self.payload)
        return _FakeExecuteResponse({})


class _FakeDb:
    def __init__(self, preview_row):
        self.preview_row = preview_row
        self.inserted: list[dict] = []

    def table(self, name: str):
        if name == "ngram_ai_preview_runs":
            return _FakePreviewBuilder(self.preview_row)
        if name == "ngram_ai_override_runs":
            return _FakeInsertBuilder(self.inserted)
        raise AssertionError(name)


def test_persist_ai_override_capture_uses_python_client_insert_shape():
    campaign_name = "Screen Shine - Duo | SPA | Cls. | Rsrch"
    workbook_path = build_workbook(
        [
            {
                "campaign_name": campaign_name,
                "category_raw": "Duo",
                "category_key": "duo",
                "mono": _make_ngram_df([]),
                "bi": _make_ngram_df(["laptop cloth"]),
                "tri": _make_ngram_df([]),
                "raw": pd.DataFrame(
                    [
                        {
                            "Search Term": "laptop cloth",
                            "Impression": 100,
                            "Click": 10,
                            "Spend": 8.5,
                            "Order 14d": 0,
                            "Sales 14d": 0,
                            "NE/NP": "",
                            "Comments": "",
                        }
                    ]
                ),
                "notes": [],
            }
        ],
        "test-version",
        ai_term_reviews={
            campaign_name: {
                "laptop cloth": {
                    "recommendation": "NEGATE",
                    "confidence": "HIGH",
                    "reason_tag": "accessory_only_intent",
                }
            }
        },
        ai_summary={
            "preview_run_id": "preview-run-456",
            "model": "gpt-5.4-2026-03-05",
            "prompt_version": "ngram_step3_calibrated_v2026_03_30",
            "spend_threshold": "1.0",
        },
    )

    try:
        reviewed_wb = load_workbook(workbook_path, data_only=True)
        fake_db = _FakeDb(
            {
                "id": "preview-run-456",
                "profile_id": "profile-1",
                "model": "gpt-5.4-2026-03-05",
                "prompt_version": "ngram_step3_calibrated_v2026_03_30",
                "preview_payload": {
                    "campaigns": [
                        {
                            "campaignName": campaign_name,
                            "synthesizedPrefills": {"mono": [], "bi": [{"gram": "laptop cloth"}], "tri": []},
                            "evaluations": [
                                {
                                    "search_term": "laptop cloth",
                                    "recommendation": "NEGATE",
                                    "confidence": "HIGH",
                                    "reason_tag": "accessory_only_intent",
                                }
                            ],
                        }
                    ]
                },
            }
        )

        result = persist_ai_override_capture(
            fake_db,
            reviewed_wb,
            collected_by_auth_user_id="user-123",
            source_filename="reviewed.xlsx",
        )

        assert result is not None
        assert result["preview_run_id"] == "preview-run-456"
        assert fake_db.inserted
        assert fake_db.inserted[0]["prompt_version"] == "ngram_step3_calibrated_v2026_03_30"
    finally:
        os.unlink(workbook_path)


def test_build_override_payload_treats_direct_negate_prefills_as_ai_prefills():
    campaign_name = "Screen Shine - Duo | SPA | Cls. | Rsrch"
    workbook_path = build_workbook(
        [
            {
                "campaign_name": campaign_name,
                "category_raw": "Duo",
                "category_key": "duo",
                "mono": _make_ngram_df([]),
                "bi": _make_ngram_df(["laptop cloth"]),
                "tri": _make_ngram_df([]),
                "raw": pd.DataFrame(
                    [
                        {
                            "Search Term": "laptop cloth",
                            "Impression": 100,
                            "Click": 10,
                            "Spend": 8.5,
                            "Order 14d": 0,
                            "Sales 14d": 0,
                            "NE/NP": "",
                            "Comments": "",
                        }
                    ]
                ),
                "notes": [],
            }
        ],
        "test-version",
        ai_term_reviews={
            campaign_name: {
                "laptop cloth": {
                    "recommendation": "NEGATE",
                    "confidence": "HIGH",
                    "reason_tag": "accessory_only_intent",
                }
            }
        },
        ai_summary={
            "preview_run_id": "preview-run-789",
            "model": "gpt-5.4-2026-03-05",
            "prompt_version": "ngram_step3_calibrated_v2026_03_30",
            "spend_threshold": "1.0",
        },
    )

    try:
        reviewed_wb = load_workbook(workbook_path, data_only=True)
        payload = _build_override_payload(
            reviewed_wb,
            {
                "id": "preview-run-789",
                "profile_id": "profile-1",
                "model": "gpt-5.4-2026-03-05",
                "prompt_version": "ngram_step3_calibrated_v2026_03_30",
                "preview_payload": {
                    "campaigns": [
                        {
                            "campaignName": campaign_name,
                            "synthesizedPrefills": {"mono": [], "bi": [], "tri": []},
                            "evaluations": [
                                {
                                    "search_term": "laptop cloth",
                                    "recommendation": "NEGATE",
                                    "confidence": "HIGH",
                                    "reason_tag": "accessory_only_intent",
                                }
                            ],
                        }
                    ]
                },
            },
        )

        assert payload is not None
        bi_diff = next(item for item in payload["campaigns"][0]["gram_diffs"]["bi"] if item["gram"] == "laptop cloth")
        assert bi_diff["ai_present"] is True
        assert bi_diff["analyst_present"] is True
        assert bi_diff["status"] == "matched"
    finally:
        os.unlink(workbook_path)
