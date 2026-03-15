import os

from openpyxl import load_workbook

from app.services.wbr.workbook import build_wbr_workbook


def test_build_wbr_workbook_creates_expected_sheets_and_headers():
    section1_report = {
        "weeks": [
            {"start": "2026-03-02", "end": "2026-03-08", "label": "02-Mar to 08-Mar"},
            {"start": "2026-02-23", "end": "2026-03-01", "label": "23-Feb to 01-Mar"},
        ],
        "rows": [
            {
                "id": "parent-1",
                "row_label": "Screen Shine | Pro",
                "row_kind": "parent",
                "parent_row_id": None,
                "sort_order": 0,
                "weeks": [
                    {"page_views": 1000, "unit_sales": 100, "sales": "2000.00", "conversion_rate": 0.1},
                    {"page_views": 900, "unit_sales": 90, "sales": "1800.00", "conversion_rate": 0.1},
                ],
            },
            {
                "id": "leaf-1",
                "row_label": "Screen Shine | Pro 2",
                "row_kind": "leaf",
                "parent_row_id": "parent-1",
                "sort_order": 1,
                "weeks": [
                    {"page_views": 400, "unit_sales": 40, "sales": "800.00", "conversion_rate": 0.1},
                    {"page_views": 350, "unit_sales": 35, "sales": "700.00", "conversion_rate": 0.1},
                ],
            },
        ],
        "qa": {},
    }
    section2_report = {
        "weeks": section1_report["weeks"],
        "rows": [
            {
                "id": "parent-1",
                "row_label": "Screen Shine | Pro",
                "row_kind": "parent",
                "parent_row_id": None,
                "sort_order": 0,
                "weeks": [
                    {
                        "impressions": 10000,
                        "clicks": 500,
                        "ctr_pct": 0.05,
                        "ad_spend": "250.00",
                        "cpc": "0.50",
                        "ad_orders": 50,
                        "ad_conversion_rate": 0.1,
                        "ad_sales": "1200.00",
                        "acos_pct": 0.2083,
                        "business_sales": "2000.00",
                        "tacos_pct": 0.125,
                    },
                    {
                        "impressions": 9000,
                        "clicks": 450,
                        "ctr_pct": 0.05,
                        "ad_spend": "225.00",
                        "cpc": "0.50",
                        "ad_orders": 45,
                        "ad_conversion_rate": 0.1,
                        "ad_sales": "1100.00",
                        "acos_pct": 0.2045,
                        "business_sales": "1800.00",
                        "tacos_pct": 0.125,
                    },
                ],
            }
        ],
        "qa": {},
    }
    section3_report = {
        "weeks": section1_report["weeks"],
        "returns_weeks": section1_report["weeks"],
        "rows": [
            {
                "id": "parent-1",
                "row_label": "Screen Shine | Pro",
                "row_kind": "parent",
                "parent_row_id": None,
                "sort_order": 0,
                "instock": 9598,
                "working": 2160,
                "reserved_plus_fc_transfer": 7388,
                "receiving_plus_intransit": 676,
                "weeks_of_stock": 11,
                "returns_week_1": 21,
                "returns_week_2": 19,
                "return_rate": 0.012,
                "_unit_sales_4w": 400,
                "_unit_sales_2w": 200,
            }
        ],
        "totals": {},
        "qa": {},
    }

    path, filename = build_wbr_workbook(
        section1_report,
        section2_report,
        section3_report,
        profile_display_name="Whoosh",
        marketplace_code="US",
        hide_empty_rows=False,
        newest_first=True,
    )

    assert os.path.exists(path)
    assert filename == "whoosh-us-wbr.xlsx"

    workbook = load_workbook(path, data_only=True)
    assert workbook.sheetnames == ["Traffic + Sales", "Advertising", "Inventory + Returns"]

    traffic_sheet = workbook["Traffic + Sales"]
    assert traffic_sheet["A1"].value == "Style"
    assert traffic_sheet["B1"].value == "Page Views"
    assert traffic_sheet.freeze_panes == "B3"

    ads_sheet = workbook["Advertising"]
    assert ads_sheet["A1"].value == "Style"
    assert ads_sheet["B1"].value == "Impressions"
    assert ads_sheet.freeze_panes == "B3"

    inventory_sheet = workbook["Inventory + Returns"]
    assert inventory_sheet["A1"].value == "Style"
    assert inventory_sheet["B1"].value == "Instock"
    assert inventory_sheet.freeze_panes == "B2"

    workbook.close()
    os.remove(path)
