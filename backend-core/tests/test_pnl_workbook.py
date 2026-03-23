from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.services.pnl.workbook import build_pnl_workbook, build_pnl_yoy_workbook


def _find_row(worksheet, label: str) -> int:
    for row_index in range(1, worksheet.max_row + 1):
        if worksheet.cell(row=row_index, column=1).value == label:
            return row_index
    raise AssertionError(f"Row not found: {label}")


def test_build_pnl_workbook_generates_dollar_and_percent_tabs(tmp_path: Path):
    report = {
        "profile": {
            "id": "p1",
            "client_id": "c1",
            "marketplace_code": "CA",
            "currency_code": "USD",
        },
        "months": ["2026-01-01", "2026-02-01"],
        "line_items": [
            {
                "key": "product_sales",
                "label": "Product Sales",
                "category": "revenue",
                "is_derived": False,
                "months": {"2026-01-01": "100.00", "2026-02-01": "200.00"},
            },
            {
                "key": "total_gross_revenue",
                "label": "Total Gross Revenue",
                "category": "summary",
                "is_derived": True,
                "months": {"2026-01-01": "100.00", "2026-02-01": "200.00"},
            },
            {
                "key": "total_refunds",
                "label": "Total Refunds & Adjustments",
                "category": "summary",
                "is_derived": True,
                "months": {"2026-01-01": "-10.00", "2026-02-01": "-20.00"},
            },
            {
                "key": "total_net_revenue",
                "label": "Total Net Revenue",
                "category": "summary",
                "is_derived": True,
                "months": {"2026-01-01": "90.00", "2026-02-01": "180.00"},
            },
            {
                "key": "cogs",
                "label": "Cost of Goods Sold",
                "category": "cogs",
                "is_derived": False,
                "months": {"2026-01-01": "30.00", "2026-02-01": "60.00"},
            },
            {
                "key": "gross_profit",
                "label": "Gross Profit",
                "category": "summary",
                "is_derived": True,
                "months": {"2026-01-01": "60.00", "2026-02-01": "120.00"},
            },
            {
                "key": "referral_fees",
                "label": "Referral Fees",
                "category": "expenses",
                "is_derived": False,
                "months": {"2026-01-01": "-15.00", "2026-02-01": "-25.00"},
            },
            {
                "key": "total_expenses",
                "label": "Total Expenses",
                "category": "summary",
                "is_derived": True,
                "months": {"2026-01-01": "-15.00", "2026-02-01": "-25.00"},
            },
            {
                "key": "net_earnings",
                "label": "Net Earnings",
                "category": "bottom_line",
                "is_derived": True,
                "months": {"2026-01-01": "45.00", "2026-02-01": "95.00"},
            },
            {
                "key": "payout_amount",
                "label": "Payout ($)",
                "category": "summary",
                "is_derived": True,
                "months": {"2026-01-01": "50.00", "2026-02-01": "90.00"},
            },
        ],
        "warnings": [],
    }

    workbook_path, filename = build_pnl_workbook(
        report,
        profile_display_name="Distex",
        marketplace_code="CA",
        currency_code="USD",
        show_totals=True,
    )

    try:
        assert filename == "distex-ca-pnl-jan2026-feb2026.xlsx"
        workbook = load_workbook(workbook_path, data_only=False)
        assert workbook.sheetnames == ["Dollars", "% of Revenue"]

        dollars_sheet = workbook["Dollars"]
        percent_sheet = workbook["% of Revenue"]

        assert dollars_sheet.cell(row=3, column=3).value == "Currency"
        assert dollars_sheet.cell(row=3, column=4).value == "CAD"
        assert dollars_sheet.cell(row=5, column=4).value == "Total"

        cogs_row = _find_row(dollars_sheet, "Cost of Goods Sold")
        assert dollars_sheet.cell(row=cogs_row, column=2).value == -30
        assert dollars_sheet.cell(row=cogs_row, column=4).value == -90

        net_margin_row = _find_row(percent_sheet, "Net Margin (%)")
        assert round(float(percent_sheet.cell(row=net_margin_row, column=2).value), 4) == 0.5

        payout_row = _find_row(dollars_sheet, "Payout ($)")
        assert dollars_sheet.cell(row=payout_row - 1, column=1).value is None
        assert dollars_sheet.cell(row=payout_row, column=2).value == 50
        assert dollars_sheet.cell(row=payout_row, column=4).value == 140

        payout_percent_row = _find_row(percent_sheet, "Payout (%)")
        assert round(float(percent_sheet.cell(row=payout_percent_row, column=2).value), 4) == round(50 / 90, 4)

        referral_fees_row = _find_row(percent_sheet, "Referral Fees")
        assert round(float(percent_sheet.cell(row=referral_fees_row, column=2).value), 4) == round(-15 / 90, 4)
        assert "[Red]" not in percent_sheet.cell(row=referral_fees_row, column=2).number_format
        assert "(" in percent_sheet.cell(row=referral_fees_row, column=2).number_format

        total_refunds_row = _find_row(percent_sheet, "Total Refunds & Adjustments")
        assert round(float(percent_sheet.cell(row=total_refunds_row, column=2).value), 4) == round(-10 / 100, 4)
        total_refunds_dollars_row = _find_row(dollars_sheet, "Total Refunds & Adjustments")
        assert "C$" in dollars_sheet.cell(row=total_refunds_dollars_row, column=2).number_format
        assert "[Red]" not in dollars_sheet.cell(row=total_refunds_dollars_row, column=2).number_format
        assert "(" in dollars_sheet.cell(row=total_refunds_dollars_row, column=2).number_format

        product_sales_row = _find_row(percent_sheet, "Product Sales")
        assert percent_sheet.cell(row=product_sales_row, column=2).value == 100
    finally:
        Path(workbook_path).unlink(missing_ok=True)


def test_build_pnl_yoy_workbook_generates_yoy_tabs(tmp_path: Path):
    report = {
        "profile": {
            "id": "p1",
            "client_id": "c1",
            "marketplace_code": "US",
            "currency_code": "USD",
        },
        "current_year": 2026,
        "prior_year": 2025,
        "months": ["Jan", "Feb"],
        "current_month_keys": ["2026-01-01", "2026-02-01"],
        "prior_month_keys": ["2025-01-01", "2025-02-01"],
        "line_items": [
            {
                "key": "product_sales",
                "label": "Product Sales",
                "category": "revenue",
                "is_derived": False,
                "current": {"2026-01-01": "110.00", "2026-02-01": "140.00"},
                "prior": {"2025-01-01": "100.00", "2025-02-01": "120.00"},
            },
            {
                "key": "total_gross_revenue",
                "label": "Total Gross Revenue",
                "category": "summary",
                "is_derived": True,
                "current": {"2026-01-01": "110.00", "2026-02-01": "140.00"},
                "prior": {"2025-01-01": "100.00", "2025-02-01": "120.00"},
            },
            {
                "key": "total_net_revenue",
                "label": "Total Net Revenue",
                "category": "summary",
                "is_derived": True,
                "current": {"2026-01-01": "100.00", "2026-02-01": "120.00"},
                "prior": {"2025-01-01": "90.00", "2025-02-01": "100.00"},
            },
            {
                "key": "advertising",
                "label": "Advertising",
                "category": "expenses",
                "is_derived": False,
                "current": {"2026-01-01": "-20.00", "2026-02-01": "-24.00"},
                "prior": {"2025-01-01": "-18.00", "2025-02-01": "-20.00"},
            },
            {
                "key": "total_expenses",
                "label": "Total Expenses",
                "category": "summary",
                "is_derived": True,
                "current": {"2026-01-01": "-30.00", "2026-02-01": "-36.00"},
                "prior": {"2025-01-01": "-27.00", "2025-02-01": "-32.00"},
            },
            {
                "key": "net_earnings",
                "label": "Net Earnings",
                "category": "bottom_line",
                "is_derived": True,
                "current": {"2026-01-01": "20.00", "2026-02-01": "24.00"},
                "prior": {"2025-01-01": "18.00", "2025-02-01": "20.00"},
            },
        ],
        "warnings": [],
        "periods": {
            "current_start": "2026-01-01",
            "current_end": "2026-02-01",
            "prior_start": "2025-01-01",
            "prior_end": "2025-02-01",
        },
    }

    workbook_path, filename = build_pnl_yoy_workbook(
        report,
        profile_display_name="Whoosh",
        marketplace_code="US",
        currency_code="USD",
        show_totals=True,
    )

    try:
        assert filename == "whoosh-us-pnl-yoy-2026.xlsx"
        workbook = load_workbook(workbook_path, data_only=False)
        assert workbook.sheetnames == ["YoY Dollars", "YoY % of Revenue"]

        dollars_sheet = workbook["YoY Dollars"]
        percent_sheet = workbook["YoY % of Revenue"]

        assert dollars_sheet.cell(row=3, column=2).value == "2026 vs 2025 (Jan - Feb)"
        product_sales_row = _find_row(dollars_sheet, "Product Sales")
        assert dollars_sheet.cell(row=product_sales_row + 1, column=1).value == 2026
        assert dollars_sheet.cell(row=product_sales_row + 1, column=2).value == 110
        assert dollars_sheet.cell(row=product_sales_row + 2, column=1).value == 2025
        assert dollars_sheet.cell(row=product_sales_row + 3, column=1).value == "Δ"
        assert round(float(dollars_sheet.cell(row=product_sales_row + 3, column=2).value), 4) == 0.1

        advertising_row = _find_row(percent_sheet, "Advertising")
        assert round(float(percent_sheet.cell(row=advertising_row + 1, column=2).value), 4) == -0.2
        assert round(float(percent_sheet.cell(row=advertising_row + 3, column=2).value), 4) == 0
    finally:
        Path(workbook_path).unlink(missing_ok=True)
