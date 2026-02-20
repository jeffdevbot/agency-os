import os
from datetime import datetime, timezone

import pandas as pd
from openpyxl import load_workbook

from app.services.root import (
    read_campaign_report,
    calculate_week_buckets,
    aggregate_hierarchy,
    build_root_workbook,
)


def test_parser_detects_currency_and_parts():
    csv_data = """Time,CampaignName,ProfileName,PortfolioName,Impression,Click,Spend,Order14d,SaleUnits14d,Sales14d
2025-12-05,Port A | SPA | MKW | Br.M | 0 - gen,Framelane [DE],Port A,10,2,£12.50,1,1,25
2025-12-04,Port B | SPM | PT | Ex. | Main,Framelane [ES],Port B,5,1,$3.00,0,0,0
"""
    df, symbol = read_campaign_report("sample.csv", csv_data.encode("utf-8"))

    # Most frequent symbol should be £ from the first row
    assert symbol == "£"
    # Parsed hierarchy parts
    assert set(df["AdType"].unique()) == {"SPA", "SPM"}
    assert set(df["Targeting"].unique()) == {"MKW", "PT"}
    assert set(df["SubType"].unique()) == {"Br.M", "Ex."}
    assert set(df["Variant"].unique()) == {"0 - gen", "Main"}


def test_week_buckets_anchor_and_labels():
    max_date = datetime(2025, 12, 6, 12, 0, tzinfo=timezone.utc)  # Saturday
    buckets = calculate_week_buckets(max_date)
    assert len(buckets) == 4
    # Week 1 should end on Dec 6 and start on Nov 30 (Sunday)
    assert buckets[0].end.date().isoformat() == "2025-12-06"
    assert buckets[0].start.date().isoformat() == "2025-11-30"
    # Labels should match dates
    assert buckets[0].start_label == "Nov 30"
    assert buckets[0].end_label == "Dec 06"


def test_aggregate_parent_first_order():
    buckets = calculate_week_buckets(datetime(2025, 12, 6, tzinfo=timezone.utc))
    df = pd.DataFrame(
        [
            {
                "ProfileName": "Framelane [DE]",
                "PortfolioName": "P1",
                "AdType": "SPA",
                "Targeting": "MKW",
                "SubType": "Br.M",
                "Variant": "0 - gen",
                "Impression": 10,
                "Click": 2,
                "Spend": 3.0,
                "Order14d": 1,
                "SaleUnits14d": 1,
                "Sales14d": 5.0,
                "Time": datetime(2025, 12, 2, tzinfo=timezone.utc),
            },
            {
                "ProfileName": "Framelane [DE]",
                "PortfolioName": "P2",
                "AdType": "SPM",
                "Targeting": "PT",
                "SubType": "Ex.",
                "Variant": "Main",
                "Impression": 4,
                "Click": 1,
                "Spend": 1.0,
                "Order14d": 0,
                "SaleUnits14d": 0,
                "Sales14d": 0.0,
                "Time": datetime(2025, 12, 1, tzinfo=timezone.utc),
            },
            {
                "ProfileName": "Framelane [ES]",
                "PortfolioName": "P3",
                "AdType": "SPA",
                "Targeting": "MKW",
                "SubType": "Br.M",
                "Variant": "1 - black",
                "Impression": 8,
                "Click": 1,
                "Spend": 2.0,
                "Order14d": 0,
                "SaleUnits14d": 0,
                "Sales14d": 0.0,
                "Time": datetime(2025, 12, 3, tzinfo=timezone.utc),
            },
        ]
    )

    nodes = aggregate_hierarchy(df, buckets)
    labels = [" | ".join(n.full_path) for n in nodes]

    # Parent-first DFS: DE tree first, then ES
    assert labels[:6] == [
        "Framelane [DE]",
        "Framelane [DE] | P1",
        "Framelane [DE] | P1 | SPA",
        "Framelane [DE] | P1 | SPA | MKW",
        "Framelane [DE] | P1 | SPA | MKW | Br.M",
        "Framelane [DE] | P1 | SPA | MKW | Br.M | 0 - gen",
    ]
    # ES root must appear after all DE nodes (DFS order) — ES subtree has 6 levels
    es_root_idx = labels.index("Framelane [ES]")
    assert es_root_idx >= 6  # after entire DE subtree


def test_workbook_builds_file(tmp_path):
    buckets = calculate_week_buckets(datetime(2025, 12, 6, tzinfo=timezone.utc))
    df = pd.DataFrame(
        [
            {
                "ProfileName": "Framelane [DE]",
                "PortfolioName": "P1",
                "AdType": "SPA",
                "Targeting": "MKW",
                "SubType": "Br.M",
                "Variant": "0 - gen",
                "Impression": 10,
                "Click": 2,
                "Spend": 3.0,
                "Order14d": 1,
                "SaleUnits14d": 1,
                "Sales14d": 5.0,
                "Time": datetime(2025, 12, 2, tzinfo=timezone.utc),
            }
        ]
    )
    nodes = aggregate_hierarchy(df, buckets)
    path = build_root_workbook(nodes, buckets, "€")
    assert os.path.exists(path)
    wb = load_workbook(path, data_only=True)
    assert "Root Keywords" in wb.sheetnames
    ws = wb["Root Keywords"]
    assert ws["A1"].value == "Hierarchy"
    wb.close()
    os.remove(path)
