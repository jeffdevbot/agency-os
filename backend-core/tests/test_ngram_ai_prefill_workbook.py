from __future__ import annotations

import os

import pandas as pd
from openpyxl import load_workbook

from app.services.ngram.workbook import build_workbook


def _make_ngram_df(grams: list[str]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "N-gram": gram,
                "Impression": 10,
                "Click": 2,
                "Spend": 1.5,
                "Order 14d": 0,
                "Sales 14d": 0,
                "CTR": 0.2,
                "CVR": 0.0,
                "CPC": 0.75,
                "ACOS": 0.0,
                "NE/NP": "",
                "Comments": "",
            }
            for gram in grams
        ]
    )


def test_build_workbook_writes_ai_prefills_to_scratchpad_columns():
    raw = pd.DataFrame(
        [
            {
                "Search Term": "travel size",
                "Impression": 100,
                "Click": 10,
                "Spend": 8.5,
                "Order 14d": 0,
                "Sales 14d": 0,
                "NE/NP": "",
                "Comments": "",
            },
            {
                "Search Term": "travel size screen cleaner",
                "Impression": 80,
                "Click": 8,
                "Spend": 6.2,
                "Order 14d": 0,
                "Sales 14d": 0,
                "NE/NP": "",
                "Comments": "",
            }
        ]
    )

    campaign_items = [
        {
            "campaign_name": "Screen Shine - Pro | SPM | MKW | Br.M | 2 - computer | Perf",
            "category_raw": "Pro",
            "category_key": "pro",
            "mono": _make_ngram_df(["travel", "size"]),
            "bi": _make_ngram_df(["travel size"]),
            "tri": _make_ngram_df(["travel size screen"]),
            "raw": raw,
            "notes": [],
        }
    ]

    workbook_path = build_workbook(
        campaign_items,
        "test-version",
        ai_term_reviews={
            "Screen Shine - Pro | SPM | MKW | Br.M | 2 - computer | Perf": {
                "travel size": {
                    "recommendation": "NEGATE",
                    "confidence": "HIGH",
                    "reason_tag": "wrong_size_variant",
                },
                "travel size screen cleaner": {
                    "recommendation": "NEGATE",
                    "confidence": "MEDIUM",
                    "reason_tag": "ambiguous_intent",
                }
            }
        },
        ai_summary={
            "preview_run_id": "preview-run-123",
            "model": "gpt-5.4-mini-2026-03-17",
            "prompt_version": "ngram_step3_calibrated_v2026_03_30",
            "spend_threshold": "1.0",
        },
    )

    try:
        wb = load_workbook(workbook_path, data_only=False)
        summary_ws = wb["Summary"]
        ws = wb[wb.sheetnames[1]]

        assert summary_ws["J3"].value == "AI Preview Run: preview-run-123"
        assert summary_ws["J4"].value == "AI Model: gpt-5.4-mini-2026-03-17"
        assert summary_ws["J5"].value == "AI Prompt Version: ngram_step3_calibrated_v2026_03_30"
        assert summary_ws["J6"].value == "AI Threshold: 1.0"

        assert ws["AV6"].value == "AI Recommendation"
        assert ws["AW6"].value == "AI Confidence"
        assert ws["AX6"].value == "AI Reason"
        assert ws["AV7"].value == "LIKELY NEGATE"
        assert ws["AW7"].value == "HIGH"
        assert ws["AX7"].value == "wrong_size_variant"
        assert ws["AV8"].value == "LIKELY NEGATE"
        assert ws["AW8"].value == "MEDIUM"
        assert ws["AX8"].value == "ambiguous_intent"
        assert ws["AT8"].value == "NE"
        assert ws["AV7"].fill.fgColor.rgb == "FFFDECEA"

        assert ws["AZ6"].value == "Monogram"
        assert ws["BA6"].value == "Bigram"
        assert ws["BB6"].value == "Trigram"
        assert ws["AY6"].value in (None, "")
        assert ws["BA7"].value == "travel size"

        assert isinstance(ws["K7"].value, str)
        assert "MATCH(A7,AZ:AZ,0)" in ws["K7"].value
        assert isinstance(ws["X7"].value, str)
        assert "MATCH(N7,BA:BA,0)" in ws["X7"].value
    finally:
        os.unlink(workbook_path)


def test_build_workbook_uses_explicit_exact_prefills_when_present():
    raw = pd.DataFrame(
        [
            {
                "Search Term": "portable monitor travel case",
                "Impression": 100,
                "Click": 10,
                "Spend": 8.5,
                "Order 14d": 0,
                "Sales 14d": 0,
                "NE/NP": "",
                "Comments": "",
            },
            {
                "Search Term": "travel screen protector bundle",
                "Impression": 80,
                "Click": 8,
                "Spend": 6.2,
                "Order 14d": 0,
                "Sales 14d": 0,
                "NE/NP": "",
                "Comments": "",
            },
        ]
    )

    campaign_name = "Screen Shine - Pro | SPM | MKW | Br.M | 2 - computer | Perf"
    workbook_path = build_workbook(
        [
            {
                "campaign_name": campaign_name,
                "category_raw": "Pro",
                "category_key": "pro",
                "mono": _make_ngram_df(["monitor"]),
                "bi": _make_ngram_df(["screen protector"]),
                "tri": _make_ngram_df(["travel monitor case"]),
                "raw": raw,
                "notes": [],
            }
        ],
        "test-version",
        ai_prefills={
            campaign_name: {
                "exact": ["portable monitor travel case"],
                "mono": [],
                "bi": ["screen protector"],
                "tri": [],
            }
        },
        ai_term_reviews={
            campaign_name: {
                "portable monitor travel case": {
                    "recommendation": "NEGATE",
                    "confidence": "HIGH",
                    "reason_tag": "accessory_only_intent",
                },
                "travel screen protector bundle": {
                    "recommendation": "NEGATE",
                    "confidence": "HIGH",
                    "reason_tag": "wrong_product_form",
                },
            }
        },
    )

    try:
        wb = load_workbook(workbook_path, data_only=False)
        ws = wb[wb.sheetnames[1]]

        assert ws["AT7"].value == "NE"
        assert ws["AT8"].value in (None, "")
        assert ws["BA7"].value == "screen protector"
        assert ws["AV7"].value == "LIKELY NEGATE"
    finally:
        os.unlink(workbook_path)


def test_build_workbook_displays_triage_labels_and_colors():
    campaign_name = "Screen Shine - Duo | SPM | MKW | Br.M | 9 - laptop | Perf"
    workbook_path = build_workbook(
        [
            {
                "campaign_name": campaign_name,
                "category_raw": "Duo",
                "category_key": "duo",
                "mono": _make_ngram_df(["juice"]),
                "bi": _make_ngram_df([]),
                "tri": _make_ngram_df([]),
                "raw": pd.DataFrame(
                    [
                        {
                            "Search Term": "computer cleaning kit",
                            "Impression": 100,
                            "Click": 10,
                            "Spend": 8.5,
                            "Order 14d": 1,
                            "Sales 14d": 10,
                            "NE/NP": "",
                            "Comments": "",
                        },
                        {
                            "Search Term": "apple juice screen cleaner",
                            "Impression": 80,
                            "Click": 8,
                            "Spend": 6.2,
                            "Order 14d": 0,
                            "Sales 14d": 0,
                            "NE/NP": "",
                            "Comments": "",
                        },
                        {
                            "Search Term": "clean macbook",
                            "Impression": 60,
                            "Click": 5,
                            "Spend": 4.1,
                            "Order 14d": 0,
                            "Sales 14d": 0,
                            "NE/NP": "",
                            "Comments": "",
                        },
                    ]
                ),
                "notes": [],
            }
        ],
        "test-version",
        ai_term_reviews={
            campaign_name: {
                "computer cleaning kit": {
                    "recommendation": "KEEP",
                    "confidence": "HIGH",
                    "reason_tag": "core_use_case",
                },
                "apple juice screen cleaner": {
                    "recommendation": "NEGATE",
                    "confidence": "HIGH",
                    "reason_tag": "wrong_category",
                },
                "clean macbook": {
                    "recommendation": "REVIEW",
                    "confidence": "MEDIUM",
                    "reason_tag": "ambiguous_intent",
                },
            }
        },
    )

    try:
        wb = load_workbook(workbook_path, data_only=False)
        ws = wb[wb.sheetnames[1]]

        assert ws["X7"].value == "SAFE KEEP"
        assert ws["X8"].value == "LIKELY NEGATE"
        assert ws["X9"].value == "REVIEW"
        assert ws["X7"].fill.fgColor.rgb == "FFE8F5E9"
        assert ws["X8"].fill.fgColor.rgb == "FFFDECEA"
        assert ws["X9"].fill.fgColor.rgb == "FFFFF4CC"
    finally:
        os.unlink(workbook_path)


def test_build_workbook_triage_mode_leaves_ne_and_scratchpad_blank():
    campaign_name = "Screen Shine - Pro | SPM | MKW | Br.M | 2 - computer | Perf"
    workbook_path = build_workbook(
        [
            {
                "campaign_name": campaign_name,
                "category_raw": "Pro",
                "category_key": "pro",
                "mono": _make_ngram_df(["travel"]),
                "bi": _make_ngram_df(["travel size"]),
                "tri": _make_ngram_df(["travel size screen"]),
                "raw": pd.DataFrame(
                    [
                        {
                            "Search Term": "travel size",
                            "Impression": 100,
                            "Click": 10,
                            "Spend": 8.5,
                            "Order 14d": 0,
                            "Sales 14d": 0,
                            "NE/NP": "",
                            "Comments": "",
                        },
                        {
                            "Search Term": "travel size screen cleaner",
                            "Impression": 80,
                            "Click": 8,
                            "Spend": 6.2,
                            "Order 14d": 0,
                            "Sales 14d": 0,
                            "NE/NP": "",
                            "Comments": "",
                        },
                    ]
                ),
                "notes": [],
            }
        ],
        "test-version",
        ai_prefills={
            campaign_name: {
                "exact": ["travel size screen cleaner"],
                "mono": ["travel"],
                "bi": ["travel size"],
                "tri": ["travel size screen"],
            }
        },
        ai_term_reviews={
            campaign_name: {
                "travel size": {
                    "recommendation": "NEGATE",
                    "confidence": "HIGH",
                    "reason_tag": "wrong_size_variant",
                },
                "travel size screen cleaner": {
                    "recommendation": "NEGATE",
                    "confidence": "MEDIUM",
                    "reason_tag": "ambiguous_intent",
                },
            }
        },
        ai_summary={
            "preview_run_id": "preview-run-triage",
            "model": "gpt-5.4-mini-2026-03-17",
            "prompt_version": "ngram_step3_calibrated_v2026_03_30",
            "spend_threshold": "1.0",
            "output_mode": "triage_only",
        },
    )

    try:
        wb = load_workbook(workbook_path, data_only=False)
        ws = wb[wb.sheetnames[1]]

        assert ws["AV7"].value == "LIKELY NEGATE"
        assert ws["AV8"].value == "LIKELY NEGATE"
        assert ws["AT7"].value in (None, "")
        assert ws["AT8"].value in (None, "")
        assert ws["AZ7"].value in (None, "")
        assert ws["BA7"].value in (None, "")
        assert ws["BB7"].value in (None, "")
    finally:
        os.unlink(workbook_path)
